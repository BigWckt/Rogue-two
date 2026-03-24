"""
Enrichissement téléphonique via 118000.fr
==========================================
Source : 118000.fr (annuaire professionnel — accessible depuis ce serveur)
Input  : mbt_villes.xlsx (onglet Synthèse) — lignes sans téléphone
Output : enrichissement_google_resultats.xlsx + mbt_villes_enrichi.xlsx

Stratégie par entreprise :
  1. Recherche 118000.fr : who={nom}&where={cp}
  2. Parcours des résultats : ne retient que la carte dont le CP correspond
     (correspondance exacte, puis correspondance département si aucun exact)
  3. Si aucun résultat sur CP, retente avec le nom de ville

Checkpointing : sauvegarde CSV tous les 50 appels (reprise possible)
Mode test     : TEST_MODE=True → ne traite que 25 lignes
"""

import re
import sys
import time
import csv
import os
import random
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

INPUT_FILE       = "mbt_villes.xlsx"
OUTPUT_ENRICH    = "enrichissement_google_resultats.xlsx"
OUTPUT_UPDATED   = "mbt_villes_enrichi.xlsx"
CHECKPOINT_FILE  = "enrichissement_checkpoint.csv"

TEST_MODE        = False   # True = 25 lignes seulement pour validation
DELAY_BETWEEN    = 1.5     # secondes entre chaque entreprise (+ jitter 0-1s)
CHECKPOINT_EVERY = 50      # sauvegarder tous les N enregistrements
BATCH_SIZE       = 100     # log intermédiaire tous les N enregistrements

# Numéros récurrents identifiés comme bruit
NOISE_PHONES = {
    "0118000",     # 118000 propre numéro
    "3118",        # numéro court 118000
}

# ── Helpers ───────────────────────────────────────────────────────────────────

PHONE_RE = re.compile(
    r"(?<!\d)((?:\+33\s?(?:\(0\)\s?)?|0)[1-9](?:[\s.\-]?\d{2}){4})(?!\d)"
)


def clean_phone(p: str) -> str:
    c = re.sub(r"[\s.\-() ]", "", p)
    if c.startswith("+33") and len(c) == 12:
        c = "0" + c[3:]
    if c.startswith("33") and len(c) == 11:
        c = "0" + c[2:]
    return c if len(c) == 10 else ""


def is_valid_phone(p: str) -> bool:
    return (
        bool(p)
        and len(p) == 10
        and p[0] == "0"
        and p not in NOISE_PHONES
        and not p.startswith(
            ("0800", "0806", "0811", "0820", "0821", "0825",
             "0890", "0891", "0892", "0893", "0897", "0898", "0899")
        )
    )


def normalize_siret(raw) -> str:
    if raw is None or str(raw).strip() in ("", "nan", "None"):
        return ""
    try:
        return str(int(float(str(raw).strip()))).zfill(14)
    except (ValueError, OverflowError):
        return str(raw).strip()


# ── Session HTTP ──────────────────────────────────────────────────────────────

session = requests.Session()
session.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
})


# ── 118000.fr search ──────────────────────────────────────────────────────────

def _extract_cp_from_card(card) -> str:
    """Extrait le code postal depuis la div .address d'une carte 118000."""
    addr = card.select_one(".address")
    if not addr:
        return ""
    m = re.search(r"\b(\d{5})\b", addr.get_text())
    return m.group(1) if m else ""


def _card_matches_cp(card, cp: str) -> bool:
    """
    Vérifie si la carte correspond au code postal recherché.
    Correspondance exacte d'abord, puis même département (2 premiers chiffres).
    """
    result_cp = _extract_cp_from_card(card)
    if not result_cp or not cp:
        return False
    if result_cp == cp:
        return True
    # Même département : cp[:2] == result_cp[:2]
    return cp[:2] == result_cp[:2]


def _first_valid_phone_from_card(card) -> str:
    """Retourne le premier numéro valide du lien tel: dans la carte."""
    for a in card.find_all("a", href=re.compile(r"^tel:")):
        raw = a["href"].replace("tel:", "")
        p = clean_phone(raw)
        if is_valid_phone(p):
            return p
    return ""


def search_118000(nom: str, ville: str, cp: str) -> str:
    """
    Cherche le numéro de téléphone sur 118000.fr.
    Retourne le numéro si trouvé (str), sinon "".
    Essai 1 : who=nom, where=cp
    Essai 2 : who=nom, where=ville  (si essai 1 échoue)
    """
    queries = [(nom[:60], cp)]
    if ville:
        queries.append((nom[:60], ville))

    for who, where in queries:
        try:
            r = session.get(
                "https://www.118000.fr/search",
                params={"who": who, "where": where},
                timeout=12,
            )
        except Exception:
            continue

        if r.status_code != 200 or len(r.text) < 1000:
            continue

        soup = BeautifulSoup(r.text, "html.parser")
        cards = soup.select("section.card")

        # Priorité : carte avec CP exact ou même département
        for card in cards:
            if _card_matches_cp(card, cp):
                phone = _first_valid_phone_from_card(card)
                if phone:
                    return phone

        # Si une seule carte et pas de CP dans l'adresse, tenter quand même
        if len(cards) == 1:
            phone = _first_valid_phone_from_card(cards[0])
            if phone:
                return phone

    return ""


# ── Enrichissement d'une entreprise ──────────────────────────────────────────

def enrich_one(nom: str, ville: str, cp: str) -> tuple[str, str]:
    """
    Retourne (téléphone_trouvé, source).
    source ∈ {"118000", ""}
    """
    phone = search_118000(nom, ville, cp)
    if phone:
        return phone, "118000"
    return "", ""


# ── Checkpoint ───────────────────────────────────────────────────────────────

def load_checkpoint() -> dict[str, str]:
    """Charge les résultats déjà traités depuis le fichier checkpoint."""
    if not os.path.exists(CHECKPOINT_FILE):
        return {}
    processed = {}
    with open(CHECKPOINT_FILE, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            siret = row.get("SIRET", "").strip()
            if siret:
                processed[siret] = row.get("Téléphone_trouvé", "")
    return processed


def save_checkpoint(results: list[dict]) -> None:
    """Sauvegarde les résultats dans le fichier checkpoint (append)."""
    fieldnames = ["SIRET", "Raison_sociale", "Téléphone_trouvé", "Source"]
    file_exists = os.path.exists(CHECKPOINT_FILE)
    with open(CHECKPOINT_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerows(results)


# ── Export Excel ──────────────────────────────────────────────────────────────

def export_enrichissement(rows: list[dict], filepath: str) -> None:
    """Export du tableau de résultats d'enrichissement."""
    cols = ["SIRET", "Raison sociale", "Code postal", "Ville",
            "Téléphone_trouvé", "Source", "Statut"]
    df = pd.DataFrame(rows, columns=cols)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Enrichissement", index=False)

    wb = load_workbook(filepath)
    ws = wb["Enrichissement"]
    hdr = PatternFill("solid", fgColor="1A3A6B")
    ok  = PatternFill("solid", fgColor="D4EDDA")
    ko  = PatternFill("solid", fgColor="FFF3CD")
    alt = PatternFill("solid", fgColor="EBF3FF")
    wht = PatternFill("solid", fgColor="FFFFFF")

    for cell in ws[1]:
        cell.fill = hdr
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tel_col = cols.index("Téléphone_trouvé") + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        base = wht if row_idx % 2 == 0 else alt
        for cell in row:
            cell.fill = base
            cell.alignment = Alignment(vertical="center")
        tc = ws.cell(row_idx, tel_col)
        if str(tc.value or "").strip():
            tc.fill = ok
            tc.font = Font(bold=True, color="1A7A1A")
        else:
            tc.fill = ko

    for idx, w in enumerate([18, 38, 12, 22, 15, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(filepath)


def update_mbt_excel(original_path: str, output_path: str,
                     phone_map: dict[str, str]) -> None:
    """
    Copie mbt_villes.xlsx et remplit la colonne Téléphone pour
    les SIRETs enrichis, dans tous les onglets.
    """
    import shutil
    shutil.copy(original_path, output_path)
    wb = load_workbook(output_path)
    ok_fill = PatternFill("solid", fgColor="D4EDDA")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, c).value or "").strip()
                   for c in range(1, ws.max_column + 1)]
        if "SIRET" not in headers or "Téléphone" not in headers:
            continue
        siret_col = headers.index("SIRET") + 1
        tel_col   = headers.index("Téléphone") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            raw_siret = str(row[siret_col - 1].value or "").strip()
            siret = normalize_siret(raw_siret)
            if siret in phone_map and phone_map[siret]:
                tel_cell = row[tel_col - 1]
                if not str(tel_cell.value or "").strip():
                    tel_cell.value = phone_map[siret]
                    tel_cell.fill  = ok_fill
                    tel_cell.font  = Font(bold=True, color="1A7A1A")

    wb.save(output_path)
    print(f"[OK] {output_path} mis à jour avec les numéros enrichis")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  ENRICHISSEMENT TÉLÉPHONIQUE — 118000.fr")
    print("  Source : mbt_villes.xlsx | Lignes sans téléphone")
    print("=" * 65)

    # ── Charger les données ───────────────────────────────────────────────────
    df = pd.read_excel(INPUT_FILE, sheet_name="Synthèse")
    df["SIRET"] = df["SIRET"].apply(normalize_siret)

    mask = df["Téléphone"].isna() | (df["Téléphone"].astype(str).str.strip() == "")
    to_enrich = df[mask].copy()

    print(f"  Total lignes fichier    : {len(df)}")
    print(f"  Lignes sans téléphone   : {len(to_enrich)}")

    if TEST_MODE:
        to_enrich = to_enrich.head(25)
        print(f"  [TEST MODE] Limité à    : {len(to_enrich)} lignes")

    # ── Charger checkpoint ────────────────────────────────────────────────────
    already_done = load_checkpoint()
    remaining    = to_enrich[~to_enrich["SIRET"].isin(already_done)].copy()
    print(f"  Déjà traités (checkpoint): {len(already_done)}")
    print(f"  À traiter               : {len(remaining)}")
    print("=" * 65)

    # ── Enrichissement ────────────────────────────────────────────────────────
    results: list[dict] = []
    batch:   list[dict] = []
    n_found = n_total = 0

    for idx, (_, row) in enumerate(remaining.iterrows(), start=1):
        siret = str(row.get("SIRET", "")).strip()
        nom   = str(row.get("Raison sociale", "")).strip()
        cp    = str(row.get("Code postal", "")).strip().replace(".0", "")
        ville = str(row.get("Ville", "")).strip()
        n_total += 1

        print(f"  [{idx:4d}/{len(remaining)}] {nom[:35]:<35} {cp}…", end=" ", flush=True)

        tel, src = enrich_one(nom, ville, cp)

        if tel:
            n_found += 1
            print(f"OK {tel} ({src})")
        else:
            print("—")

        record = {
            "SIRET":           siret,
            "Raison sociale":  nom,
            "Code postal":     cp,
            "Ville":           ville,
            "Téléphone_trouvé": tel,
            "Source":          src,
            "Statut":          "Trouvé" if tel else "Non trouvé",
        }
        results.append(record)
        batch.append({
            "SIRET": siret, "Raison_sociale": nom,
            "Téléphone_trouvé": tel, "Source": src,
        })

        # Checkpoint
        if len(batch) >= CHECKPOINT_EVERY:
            save_checkpoint(batch)
            batch = []
            pct = n_found / n_total * 100
            print(f"\n  ── Checkpoint {idx}/{len(remaining)} │ "
                  f"{n_found} trouvés ({pct:.1f}%) ──\n")

        # Log intermédiaire
        if n_total % BATCH_SIZE == 0 and n_total > 0:
            pct = n_found / n_total * 100
            print(f"\n  ── Bilan {n_total}/{len(remaining)} │ "
                  f"{n_found} trouvés ({pct:.1f}%) ──\n", flush=True)

        # Délai entre requêtes
        time.sleep(DELAY_BETWEEN + random.uniform(0, 1.0))

    if batch:
        save_checkpoint(batch)

    # ── Fusionner avec checkpoint précédent ───────────────────────────────────
    for siret, tel in already_done.items():
        nom_row = df[df["SIRET"] == siret]
        nom  = nom_row["Raison sociale"].iloc[0] if len(nom_row) else ""
        cp   = str(nom_row["Code postal"].iloc[0] if len(nom_row) else "").replace(".0","")
        ville= nom_row["Ville"].iloc[0] if len(nom_row) else ""
        results.append({
            "SIRET": siret, "Raison sociale": nom,
            "Code postal": cp, "Ville": ville,
            "Téléphone_trouvé": tel, "Source": "checkpoint",
            "Statut": "Trouvé" if tel else "Non trouvé",
        })

    # ── Bilan console ─────────────────────────────────────────────────────────
    all_found   = [r for r in results if r["Téléphone_trouvé"]]
    all_not     = [r for r in results if not r["Téléphone_trouvé"]]
    total_r     = len(results)
    pct_found   = len(all_found) / max(total_r, 1) * 100

    print("\n" + "=" * 65)
    print("  BILAN ENRICHISSEMENT")
    print("=" * 65)
    print(f"  SIRETs traités          : {total_r}")
    print(f"  Téléphones trouvés      : {len(all_found)}  ({pct_found:.1f}%)")
    print(f"  Non trouvés             : {len(all_not)}  ({100-pct_found:.1f}%)")
    print()
    src_counts: dict[str, int] = {}
    for r in all_found:
        s = r["Source"]
        src_counts[s] = src_counts.get(s, 0) + 1
    for src, cnt in sorted(src_counts.items(), key=lambda x: -x[1]):
        print(f"    {src:<15}: {cnt}")
    print("=" * 65)

    # ── Export Excel résultats ────────────────────────────────────────────────
    export_enrichissement(results, OUTPUT_ENRICH)
    print(f"\n[OK] Résultats exportés : {OUTPUT_ENRICH}")

    # ── Mettre à jour mbt_villes avec les numéros trouvés ─────────────────────
    phone_map = {r["SIRET"]: r["Téléphone_trouvé"]
                 for r in results if r["Téléphone_trouvé"]}
    update_mbt_excel(INPUT_FILE, OUTPUT_UPDATED, phone_map)

    print(f"\n  {len(phone_map)} numéros injectés dans {OUTPUT_UPDATED}")
    print("  Terminé\n")


if __name__ == "__main__":
    main()
