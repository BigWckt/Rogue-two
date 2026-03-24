"""
Test exploratoire de l'API Pappers v2
Objectif : mesurer le taux d'enrichissement réel pour des boulangeries IDF

RÉSULTATS PRÉLIMINAIRES (constatés lors de l'exploration de l'API) :
  - Téléphone / Email / Site web : NON DISPONIBLES via l'API Pappers
    → Pappers s'appuie exclusivement sur les registres officiels (INSEE, INPI,
      BODACC) qui ne contiennent pas ces données de contact.
  - Dirigeants : comptés (nb_dirigeants_total) mais non retournés dans /recherche
    → Pour les noms de dirigeants, il faudrait appeler /entreprise par SIRET
      (1 crédit supplémentaire par SIRET).
  - Adresse / SIRET / NAF / Statut / Effectif : 100% disponibles

Codes NAF ciblés : 10.71A, 10.71B, 10.71C, 10.71D
Zone : Île-de-France (75, 77, 78, 91, 92, 93, 94, 95)
Budget : max 50 résultats par (code_naf × département) → ~200 résultats max
Coût estimé : 200 résultats × 0.10 crédit = ~20 crédits
"""

import time
import sys
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ─────────────────────────────────────────────────────────────

API_KEY  = "04a0984d429ea90c8fe17311daf2a175338f475b9b2ee4a1"
BASE_URL = "https://api.pappers.fr/v2"

NAF_CODES = [
    ("10.71A", "Fabrication industrielle de pain et pâtisserie fraîche"),
    ("10.71B", "Cuisson de produits de boulangerie"),
    ("10.71C", "Boulangerie et boulangerie-pâtisserie artisanale"),
    ("10.71D", "Pâtisserie artisanale"),
]

DEPTS_IDF = ["75", "77", "78", "91", "92", "93", "94", "95"]

# Nombre max de résultats à récupérer au total (budget crédits limité)
MAX_PAR_NAF     = 50    # max par code NAF (toutes villes confondues)
DELAY_BETWEEN_CALLS = 0.5  # secondes entre chaque appel

OUTPUT_FILE = "test_pappers_resultats.xlsx"

# Décodage des tranches d'effectif INSEE
TRANCHE_EFFECTIF = {
    "00": "0 salarié",
    "01": "1-2 salariés",
    "02": "3-5 salariés",
    "03": "6-9 salariés",
    "11": "10-19 salariés",
    "12": "20-49 salariés",
    "21": "50-99 salariés",
    "22": "100-199 salariés",
    "31": "200-249 salariés",
    "32": "250-499 salariés",
    "41": "500-999 salariés",
    "42": "1 000-1 999 salariés",
    "51": "2 000-4 999 salariés",
    "52": "5 000-9 999 salariés",
    "53": "10 000+ salariés",
}

EXCEL_COLUMNS = [
    "Source",
    "SIRET",
    "Raison sociale",
    "Adresse",
    "Code postal",
    "Ville",
    "Téléphone",      # toujours vide — Pappers ne fournit pas cette donnée
    "Email",          # toujours vide — idem
    "Site web",       # toujours vide — idem
    "Code NAF",
    "Libellé NAF",
    "Effectif (tranche)",
    "Effectif (libellé)",
    "Nb dirigeants",  # le compte, pas les noms
    "Statut",
    "Date création",
    "Forme juridique",
    "Latitude",
    "Longitude",
]

# ── Fonctions utilitaires ─────────────────────────────────────────────────────

def http_get(url: str, params: dict, retries: int = 3) -> dict | None:
    """GET avec retry exponentiel."""
    for attempt in range(retries):
        try:
            resp = requests.get(url, params=params, timeout=20)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code == 429:
                wait = 2 ** attempt
                print(f"    ⏳ Rate limit — attente {wait}s…")
                time.sleep(wait)
                continue
            if resp.status_code == 401:
                print(f"    ❌ Clé API invalide (401)")
                sys.exit(1)
            print(f"    ⚠️  HTTP {resp.status_code}: {resp.text[:100]}")
            return None
        except requests.RequestException as exc:
            wait = 2 ** attempt
            print(f"    ⚠️  Erreur réseau ({exc}) — retry dans {wait}s")
            time.sleep(wait)
    return None


def build_row(r: dict, naf_label: str) -> dict:
    """Transforme un résultat Pappers en ligne pour le DataFrame."""
    siege = r.get("siege") or {}
    adresse_parts = [
        p for p in [
            siege.get("adresse_ligne_1"),
            siege.get("adresse_ligne_2"),
        ] if p
    ]
    tranche_code = r.get("tranche_effectif") or ""
    return {
        "Source":             "Pappers",
        "SIRET":              siege.get("siret") or "",
        "Raison sociale":     r.get("nom_entreprise") or "",
        "Adresse":            ", ".join(adresse_parts),
        "Code postal":        siege.get("code_postal") or "",
        "Ville":              siege.get("ville") or "",
        "Téléphone":          "",   # non disponible via Pappers API
        "Email":              "",   # non disponible via Pappers API
        "Site web":           "",   # non disponible via Pappers API
        "Code NAF":           r.get("code_naf") or "",
        "Libellé NAF":        naf_label,
        "Effectif (tranche)": tranche_code,
        "Effectif (libellé)": TRANCHE_EFFECTIF.get(tranche_code, r.get("effectif") or ""),
        "Nb dirigeants":      r.get("nb_dirigeants_total") or 0,
        "Statut":             r.get("statut_consolide") or (
                                  "cessé" if r.get("entreprise_cessee") else "actif"
                              ),
        "Date création":      r.get("date_creation_formate") or "",
        "Forme juridique":    r.get("forme_juridique") or "",
        "Latitude":           siege.get("latitude") or "",
        "Longitude":          siege.get("longitude") or "",
    }


# ── Collecte des données ──────────────────────────────────────────────────────

def fetch_pappers(naf_code: str, naf_label: str, max_results: int = MAX_PAR_NAF) -> list[dict]:
    """
    Interroge /recherche pour un code NAF sur tous les départements IDF.
    Arrête dès que max_results est atteint.
    Retourne une liste de lignes formatées.
    """
    rows: list[dict] = []
    seen_sirets: set[str] = set()

    dept_str = ",".join(DEPTS_IDF)

    print(f"\n  [Pappers] NAF={naf_code} — depts={dept_str}")

    page = 1
    while len(rows) < max_results:
        remaining = max_results - len(rows)
        par_page  = min(remaining, 50)   # max 50 par page selon doc Pappers

        params = {
            "api_token":        API_KEY,
            "code_naf":         naf_code,
            "departement":      dept_str,
            "par_page":         par_page,
            "page":             page,
            "entreprise_cessee": False,
        }

        data = http_get(f"{BASE_URL}/recherche", params=params)
        if data is None:
            break

        resultats = data.get("resultats") or []
        total     = data.get("total", 0)

        if not resultats:
            break

        for r in resultats:
            siret = (r.get("siege") or {}).get("siret") or r.get("siren") or ""
            if siret in seen_sirets:
                continue
            seen_sirets.add(siret)
            rows.append(build_row(r, naf_label))
            if len(rows) >= max_results:
                break

        print(f"    page {page}: {len(resultats)} résultats "
              f"(cumulé: {len(rows)}/{min(max_results, total)} — total dispo: {total})")

        if len(resultats) < par_page or len(rows) >= max_results:
            break

        page += 1
        time.sleep(DELAY_BETWEEN_CALLS)

    return rows


# ── Export Excel ──────────────────────────────────────────────────────────────

def export_excel(rows: list[dict], filepath: str) -> None:
    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    total = len(df)

    # ── Calcul des taux de remplissage ──
    def fill_rate(col: str) -> tuple[int, float]:
        n = df[col].apply(lambda x: bool(str(x).strip()) and str(x) not in ("", "0", "nan")).sum()
        return int(n), round(n / total * 100, 1) if total else 0.0

    stats = {col: fill_rate(col) for col in EXCEL_COLUMNS}

    # ── Synthèse console ──
    print("\n" + "=" * 60)
    print("  SYNTHÈSE — Résultats Pappers API")
    print("=" * 60)
    print(f"  Entreprises collectées           : {total}")
    print(f"  Actives                          : {(df['Statut'] == 'actif').sum()}")
    print()
    print("  Taux de remplissage des champs :")
    print(f"  {'Champ':<30} {'Renseigné':>10}  {'Taux':>7}")
    print("  " + "-" * 52)
    fields_to_report = [
        "SIRET", "Raison sociale", "Adresse", "Code postal", "Ville",
        "Téléphone", "Email", "Site web",
        "Code NAF", "Effectif (tranche)", "Nb dirigeants",
        "Statut", "Date création", "Forme juridique",
    ]
    for col in fields_to_report:
        n, pct = stats[col]
        warning = " ⚠️  NON DISPONIBLE VIA PAPPERS" if col in ("Téléphone", "Email", "Site web") else ""
        print(f"  {col:<30} {n:>8}    {pct:>5.1f}%{warning}")

    print()
    print("  ⚠️  CONCLUSION ENRICHISSEMENT :")
    print("  Pappers (sources: INSEE/INPI/BODACC) NE contient PAS :")
    print("  → Téléphone   (0% fill rate)")
    print("  → Email       (0% fill rate)")
    print("  → Site web    (0% fill rate)")
    print()
    print("  Pappers EST utile pour :")
    n_dir, pct_dir = stats["Nb dirigeants"]
    n_eff, pct_eff = stats["Effectif (tranche)"]
    n_adr, pct_adr = stats["Adresse"]
    print(f"  → SIRET / Adresse structurée     ({pct_adr:.0f}% fill)")
    print(f"  → Tranche d'effectif             ({pct_eff:.0f}% fill)")
    print(f"  → Nombre de dirigeants           ({pct_dir:.0f}% fill)")
    print(f"  → Vérification statut actif/cessé")
    print(f"  → Volume: {total} entreprises en IDF sur codes NAF ciblés")
    print("=" * 60)

    # ── Export Excel ──
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Pappers_test", index=False)

    wb = load_workbook(filepath)
    ws = wb["Pappers_test"]

    # En-têtes
    header_fill = PatternFill("solid", fgColor="1A3A6B")
    warn_fill   = PatternFill("solid", fgColor="C0392B")   # rouge pour champs vides
    alt_fill    = PatternFill("solid", fgColor="E8F0FF")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    # Colonnes "toujours vides" à signaler en rouge
    empty_cols = {"Téléphone", "Email", "Site web"}

    col_indices = {col: idx + 1 for idx, col in enumerate(EXCEL_COLUMNS)}

    for cell in ws[1]:
        col_name = EXCEL_COLUMNS[cell.column - 1]
        cell.fill      = warn_fill if col_name in empty_cols else header_fill
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = white_fill if row_idx % 2 == 0 else alt_fill
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")

    # Ajuster les largeurs
    col_widths = {
        "Source": 10, "SIRET": 18, "Raison sociale": 35,
        "Adresse": 40, "Code postal": 12, "Ville": 20,
        "Téléphone": 15, "Email": 25, "Site web": 25,
        "Code NAF": 12, "Libellé NAF": 38,
        "Effectif (tranche)": 10, "Effectif (libellé)": 22,
        "Nb dirigeants": 14, "Statut": 12,
        "Date création": 15, "Forme juridique": 25,
        "Latitude": 12, "Longitude": 12,
    }
    for col_name, width in col_widths.items():
        if col_name in col_indices:
            letter = ws.cell(1, col_indices[col_name]).column_letter
            ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"

    # Onglet synthèse
    ws2 = wb.create_sheet("Synthèse_enrichissement")
    synthese_data = [
        ["Champ", "Nb renseigné", "Total", "Taux (%)", "Disponible via Pappers ?"],
    ]
    availability = {
        "SIRET": "✅ Oui", "Raison sociale": "✅ Oui",
        "Adresse": "✅ Oui", "Code postal": "✅ Oui", "Ville": "✅ Oui",
        "Téléphone": "❌ Non — absent des registres INSEE/INPI",
        "Email": "❌ Non — absent des registres INSEE/INPI",
        "Site web": "❌ Non — absent des registres INSEE/INPI",
        "Code NAF": "✅ Oui", "Libellé NAF": "✅ Oui",
        "Effectif (tranche)": "✅ Oui (INSEE)", "Effectif (libellé)": "✅ Oui",
        "Nb dirigeants": "✅ Compte seulement (noms via /entreprise +1 crédit/SIRET)",
        "Statut": "✅ Oui", "Date création": "✅ Oui", "Forme juridique": "✅ Oui",
    }
    for col in fields_to_report:
        n, pct = stats[col]
        synthese_data.append([col, n, total, pct, availability.get(col, "?")])

    for row in synthese_data:
        ws2.append(row)

    # Mise en forme synthèse
    for cell in ws2[1]:
        cell.fill = PatternFill("solid", fgColor="1A3A6B")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 55

    wb.save(filepath)
    print(f"\n[OK] Fichier Excel exporté : {filepath}")
    print(f"     Onglets : 'Pappers_test' ({total} lignes) + 'Synthèse_enrichissement'")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  TEST API PAPPERS — Boulangeries Île-de-France")
    print(f"  NAF  : {', '.join(c for c, _ in NAF_CODES)}")
    print(f"  Depts: {', '.join(DEPTS_IDF)}")
    print(f"  Max  : {MAX_PAR_NAF} résultats / code NAF (~{MAX_PAR_NAF * len(NAF_CODES) * 0.1:.0f} crédits)")
    print("=" * 60)
    print()
    print("  ℹ️  NOTE : Pappers ne contient PAS de téléphone/email/site web")
    print("     (sources : INSEE, INPI, BODACC uniquement)")
    print()

    all_rows: list[dict] = []
    seen_sirets: set[str] = set()

    for naf_code, naf_label in NAF_CODES:
        rows = fetch_pappers(naf_code, naf_label)
        # Déduplication inter-NAF par SIRET
        for row in rows:
            siret = row["SIRET"]
            if siret not in seen_sirets:
                seen_sirets.add(siret)
                all_rows.append(row)
        print(f"  → {len(rows)} résultats NAF={naf_code} | "
              f"cumulé unique: {len(all_rows)}")
        time.sleep(DELAY_BETWEEN_CALLS)

    print(f"\n[INFO] Total entreprises uniques : {len(all_rows)}")

    if not all_rows:
        print("[WARN] Aucune entreprise trouvée — vérifier la clé API.")
        sys.exit(0)

    export_excel(all_rows, OUTPUT_FILE)


if __name__ == "__main__":
    main()
