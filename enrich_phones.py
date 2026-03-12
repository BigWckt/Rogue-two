"""
Enrichissement des numéros de téléphone manquants dans le fichier Excel.
Sources (dans l'ordre) :
  1. Societe.com  — recherche par SIRET, redirect vers page entreprise
  2. Pages Jaunes — recherche par nom + ville (fallback si Societe.com échoue)

Usage : python3 enrich_phones.py
"""

import re
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ─────────────────────────────────────────────────────────────

INPUT_FILE  = "entreprises_alternance_paris.xlsx"
SHEET_NAME  = "Résultats_Scrap"
DELAY_SEC   = 0.5   # délai entre chaque appel HTTP

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.google.fr/",
}

# Regex numéro FR : 10 chiffres commençant par 0[1-9]
PHONE_RE = re.compile(r'\b(0[1-9](?:[\s.\-]?\d{2}){4})\b')

# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_phone(raw: str) -> str:
    """Normalise un numéro en 10 chiffres sans séparateurs."""
    digits = re.sub(r'\D', '', raw)
    return digits if len(digits) == 10 else ""


def _extract_first_phone(html: str, exclude_digits: set[str] | None = None) -> str:
    """Extrait le premier numéro de téléphone valide du HTML."""
    exclude_digits = exclude_digits or set()
    for m in PHONE_RE.finditer(html):
        cleaned = _clean_phone(m.group(1))
        if cleaned and cleaned not in exclude_digits:
            return cleaned
    return ""


# ── Source 1 : Societe.com ────────────────────────────────────────────────────

def fetch_societe_com(siret: str) -> str:
    """Cherche un téléphone sur societe.com via SIRET."""
    if not siret:
        return ""
    try:
        url = f"https://www.societe.com/cgi-bin/search?champs={siret}"
        resp = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        if resp.status_code != 200:
            return ""
        # Vérifier qu'on est bien sur une page entreprise (pas résultats de recherche)
        if "/societe/" not in resp.url and "/personne/" not in resp.url:
            return ""
        # Exclure les chiffres du SIRET lui-même pour éviter les faux positifs
        siret_fragments = {siret[i:i+10] for i in range(len(siret)-9)}
        phone = _extract_first_phone(resp.text, exclude_digits=siret_fragments)
        return phone
    except Exception:
        return ""


# ── Source 2 : Pages Jaunes ───────────────────────────────────────────────────

def fetch_pages_jaunes(nom: str, ville: str) -> str:
    """Cherche un téléphone sur pagesjaunes.fr via nom + ville."""
    if not nom or not ville:
        return ""
    try:
        query = requests.utils.quote(nom[:50])
        loc   = requests.utils.quote(ville[:30])
        url   = f"https://www.pagesjaunes.fr/annuaire/chercherlesnoms?quoiqui={query}&ou={loc}"
        resp  = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return ""
        # Pages Jaunes masque souvent les numéros via JS — tenter extraction directe
        soup = BeautifulSoup(resp.text, "lxml")
        # Chercher data-href avec tel: ou itemprop telephone
        for tag in soup.find_all(href=re.compile(r'^tel:')):
            cleaned = _clean_phone(tag["href"].replace("tel:", ""))
            if cleaned:
                return cleaned
        for tag in soup.find_all(attrs={"itemprop": "telephone"}):
            cleaned = _clean_phone(tag.get_text())
            if cleaned:
                return cleaned
        # Fallback regex brut
        return _extract_first_phone(resp.text)
    except Exception:
        return ""


# ── Enrichissement principal ──────────────────────────────────────────────────

def enrich_excel(filepath: str) -> None:
    wb = load_workbook(filepath)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] Feuille '{SHEET_NAME}' introuvable dans {filepath}")
        return
    ws = wb[SHEET_NAME]

    # Lire l'en-tête
    headers = [cell.value for cell in ws[1]]

    def col(name: str) -> int | None:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    col_tel    = col("Téléphone")
    col_siret  = col("SIRET")
    col_nom    = col("Raison sociale")
    col_ville  = col("Ville")

    if not all([col_tel, col_siret, col_nom, col_ville]):
        print(f"[ERROR] Colonnes manquantes : Téléphone={col_tel} SIRET={col_siret}")
        return

    # Créer ou récupérer la colonne "Statut enrichissement"
    col_statut = col("Statut enrichissement")
    if col_statut is None:
        col_statut = len(headers) + 1
        ws.cell(row=1, column=col_statut, value="Statut enrichissement")
        # Appliquer le style de l'en-tête existant
        ref_cell = ws.cell(row=1, column=col_tel)
        new_header = ws.cell(row=1, column=col_statut)
        new_header.fill      = PatternFill("solid", fgColor="0066FF")
        new_header.font      = Font(bold=True, color="FFFFFF")
        new_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Parcourir les lignes de données
    nb_total      = 0
    nb_enrichis   = 0
    nb_non_trouve = 0
    nb_deja_rempli = 0

    print(f"\n[ENRICH] Démarrage — fichier : {filepath}")
    print(f"[ENRICH] Source 1 : societe.com | Source 2 : pagesjaunes.fr\n")

    for row_idx in range(2, ws.max_row + 1):
        tel_cell   = ws.cell(row=row_idx, column=col_tel)
        siret_cell = ws.cell(row=row_idx, column=col_siret)
        nom_cell   = ws.cell(row=row_idx, column=col_nom)
        ville_cell = ws.cell(row=row_idx, column=col_ville)

        # Ignorer les lignes de synthèse (cellule SIRET vide)
        siret_val = str(siret_cell.value or "").strip()
        if not siret_val or siret_val in ("None", "SIRET"):
            continue

        nb_total += 1
        tel_val = str(tel_cell.value or "").strip()

        # Ne pas écraser les téléphones existants
        if tel_val and tel_val.lower() not in ("none", "nan", ""):
            nb_deja_rempli += 1
            continue

        nom   = str(nom_cell.value   or "").strip()
        ville = str(ville_cell.value or "").strip()

        # Normaliser SIRET (peut être float '80350239200019.0')
        siret = re.sub(r'\.0$', '', siret_val).zfill(14)

        # --- Source 1 : Societe.com ---
        time.sleep(DELAY_SEC)
        phone = fetch_societe_com(siret)
        source = "societe.com"

        # --- Source 2 : Pages Jaunes (fallback) ---
        if not phone:
            time.sleep(DELAY_SEC)
            phone = fetch_pages_jaunes(nom, ville)
            source = "pagesjaunes.fr"

        statut_cell = ws.cell(row=row_idx, column=col_statut)

        if phone:
            # Formater en XX XX XX XX XX
            fmt = " ".join(phone[i:i+2] for i in range(0, 10, 2))
            tel_cell.value    = fmt
            statut_cell.value = f"Enrichi via {source}"
            nb_enrichis += 1
            print(f"  ✓ {nom[:35]:<35} → {fmt}  [{source}]")
        else:
            statut_cell.value = "Non trouvé"
            nb_non_trouve += 1
            print(f"  ✗ {nom[:35]:<35} → non trouvé")

    # ── Ajuster la largeur de la colonne Statut enrichissement ──
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(col_statut)
    ws.column_dimensions[col_letter].width = 28

    wb.save(filepath)

    print(f"""
── Synthèse enrichissement ──────────────────────────
  Lignes analysées             : {nb_total}
  Déjà renseignées (inchangées): {nb_deja_rempli}
  Téléphones récupérés         : {nb_enrichis}
  Non trouvés                  : {nb_non_trouve}
─────────────────────────────────────────────────────
[OK] Fichier sauvegardé : {filepath}
""")


# ── Point d'entrée ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    enrich_excel(INPUT_FILE)
