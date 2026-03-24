"""
Test d'enrichissement téléphonique via Pappers /v2/entreprise
Endpoint : /v2/entreprise?siret={SIRET}&champs_supplementaires=contacts_generiques

Objectif : mesurer le taux de couverture réel du champ contacts_generiques
           (téléphone + email) sur un échantillon de 100 SIRETs IDF boulangerie

Budget    : 1 crédit par appel → 100 crédits max
Sources   : test_pappers_resultats.xlsx (200 SIRETs) + entreprises_alternance_paris.xlsx (63 SIRETs)
            → prend les 100 premiers SIRETs uniques disponibles

Sortie    : logs console détaillés + test_pappers_enrichissement.xlsx
"""

import time
import sys
import re
import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────────────────

API_KEY      = "04a0984d429ea90c8fe17311daf2a175338f475b9b2ee4a1"
BASE_URL     = "https://api.pappers.fr/v2/entreprise"
MAX_SIRETS   = 100
DELAY        = 0.5   # secondes entre appels (respect rate-limit)
OUTPUT_FILE  = "test_pappers_enrichissement.xlsx"

SOURCE_FILES = [
    "test_pappers_resultats.xlsx",
    "entreprises_alternance_paris.xlsx",
]

# ── Collecte des SIRETs source ────────────────────────────────────────────────

def load_sirets() -> list[tuple[str, str, str]]:
    """
    Charge jusqu'à MAX_SIRETS SIRETs uniques depuis les fichiers Excel disponibles.
    Retourne une liste de (siret, raison_sociale, source_fichier).
    """
    rows: list[tuple[str, str, str]] = []
    seen: set[str] = set()

    for fname in SOURCE_FILES:
        try:
            df = pd.read_excel(fname)
        except FileNotFoundError:
            print(f"  [WARN] Fichier non trouvé : {fname}")
            continue

        for _, row in df.iterrows():
            if len(rows) >= MAX_SIRETS:
                break
            raw_siret = str(row.get("SIRET") or "").strip().replace(" ", "")
            if not raw_siret or len(raw_siret) < 9 or raw_siret in seen:
                continue
            seen.add(raw_siret)
            nom = str(row.get("Raison sociale") or "").strip()
            rows.append((raw_siret, nom, fname))

        print(f"  {fname} : {len([r for r in rows if r[2] == fname])} SIRETs chargés")

    return rows[:MAX_SIRETS]


# ── Appel API ──────────────────────────────────────────────────────────────────

def fetch_contacts(siret: str) -> dict:
    """
    Appelle /v2/entreprise avec champs_supplementaires=contacts_generiques.
    Retourne un dict avec les champs extraits + métadonnées de la réponse.
    """
    try:
        resp = requests.get(
            BASE_URL,
            params={
                "api_token":            API_KEY,
                "siret":                siret,
                "champs_supplementaires": "contacts_generiques",
            },
            timeout=15,
        )
    except requests.RequestException as exc:
        return {"status": "ERREUR_RESEAU", "error": str(exc)}

    if resp.status_code == 401:
        print("  ❌ Clé API invalide (401) — arrêt")
        sys.exit(1)
    if resp.status_code == 429:
        return {"status": "RATE_LIMIT"}
    if resp.status_code != 200:
        return {"status": f"HTTP_{resp.status_code}", "error": resp.text[:80]}

    data = resp.json()

    # Extraire contacts_generiques
    contacts = data.get("contacts_generiques")      # attendu : list[dict] ou None

    telephone = ""
    email     = ""
    nb_contacts = 0

    if contacts and isinstance(contacts, list):
        nb_contacts = len(contacts)
        for c in contacts:
            if not telephone and c.get("telephone"):
                telephone = str(c["telephone"]).strip()
            if not email and c.get("email"):
                email = str(c["email"]).strip()
    elif contacts and isinstance(contacts, dict):
        nb_contacts = 1
        telephone = str(contacts.get("telephone") or "").strip()
        email     = str(contacts.get("email") or "").strip()

    return {
        "status":            "OK",
        "nom_entreprise":    data.get("nom_entreprise") or "",
        "statut_consolide":  data.get("statut_consolide") or "",
        "contacts_raw":      str(contacts) if contacts else "null",
        "nb_contacts":       nb_contacts,
        "telephone":         telephone,
        "email":             email,
        "champs_dispo":      sorted(data.keys()),
    }


# ── Export Excel ───────────────────────────────────────────────────────────────

RESULT_COLS = [
    "SIRET", "Raison sociale (source)", "Raison sociale (Pappers)",
    "Source fichier", "Statut Pappers",
    "contacts_generiques (brut)", "Nb contacts",
    "Téléphone", "Email",
    "Résultat appel",
]


def export_excel(results: list[dict], filepath: str) -> None:
    df = pd.DataFrame(results, columns=RESULT_COLS)
    total = len(df)

    n_tel   = df["Téléphone"].apply(lambda x: bool(str(x).strip())).sum()
    n_email = df["Email"].apply(lambda x: bool(str(x).strip())).sum()
    n_ok    = (df["Résultat appel"] == "OK").sum()
    n_err   = total - int(n_ok)
    n_null  = (df["contacts_generiques (brut)"] == "null").sum()
    n_data  = total - int(n_null)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Enrichissement_test", index=False)

    wb = load_workbook(filepath)
    ws = wb["Enrichissement_test"]

    hdr_fill  = PatternFill("solid", fgColor="1A3A6B")
    ok_fill   = PatternFill("solid", fgColor="D4EDDA")
    ko_fill   = PatternFill("solid", fgColor="F8D7DA")
    alt_fill  = PatternFill("solid", fgColor="EBF3FF")
    white_fill= PatternFill("solid", fgColor="FFFFFF")

    # En-têtes
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colonnes téléphone / email colorées si renseignées
    tel_col   = RESULT_COLS.index("Téléphone") + 1
    email_col = RESULT_COLS.index("Email") + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        base_fill = white_fill if row_idx % 2 == 0 else alt_fill
        for cell in row:
            cell.fill      = base_fill
            cell.alignment = Alignment(vertical="center")
        # Surligner téléphone / email si trouvé
        tel_cell   = ws.cell(row_idx, tel_col)
        email_cell = ws.cell(row_idx, email_col)
        if str(tel_cell.value or "").strip():
            tel_cell.fill = ok_fill
            tel_cell.font = Font(bold=True, color="1A7A1A")
        if str(email_cell.value or "").strip():
            email_cell.fill = ok_fill
            email_cell.font = Font(bold=True, color="1A7A1A")

    # Largeurs de colonnes
    for col_letter, width in zip(
        ["A","B","C","D","E","F","G","H","I","J"],
        [18, 35, 35, 30, 15, 35, 12, 18, 30, 14]
    ):
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Onglet synthèse
    ws2 = wb.create_sheet("Synthèse")
    run_date = datetime.now().strftime("%d/%m/%Y %H:%M")
    synthese = [
        ["Paramètre", "Valeur"],
        ["Date test", run_date],
        ["Endpoint", f"{BASE_URL}?champs_supplementaires=contacts_generiques"],
        ["SIRETs testés", total],
        ["Appels réussis (HTTP 200)", int(n_ok)],
        ["Appels en erreur", n_err],
        ["", ""],
        ["RÉSULTATS contacts_generiques", ""],
        ["contacts_generiques = null", int(n_null)],
        ["contacts_generiques ≠ null (data)", int(n_data)],
        ["", ""],
        ["TAUX DE COUVERTURE", ""],
        ["Avec téléphone", f"{int(n_tel)} / {total} ({n_tel/total*100:.1f}%)"],
        ["Avec email",     f"{int(n_email)} / {total} ({n_email/total*100:.1f}%)"],
        ["", ""],
        ["CONCLUSION", ""],
        ["contacts_generiques disponible ?",
         "OUI" if int(n_data) > 0 else "NON — champ toujours null (abonnement insuffisant ou donnée non collectée)"],
        ["Recommandation",
         "Champ premium non disponible — Pappers ne fournit pas de téléphone via cette API"
         if int(n_tel) == 0 else
         f"Pappers fournit le téléphone pour {n_tel/total*100:.1f}% des entreprises testées"],
    ]
    for row in synthese:
        ws2.append(row)

    # Mise en forme synthèse
    for cell in ws2[1]:
        cell.fill = PatternFill("solid", fgColor="1A3A6B")
        cell.font = Font(bold=True, color="FFFFFF")
    for row_idx in [8, 12, 16]:
        try:
            for cell in ws2[row_idx]:
                cell.font = Font(bold=True, color="1A3A6B")
                cell.fill = PatternFill("solid", fgColor="D0E4FF")
        except Exception:
            pass

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 75

    # Résultat final en rouge/vert
    conclusion_row = len(synthese) + 1
    try:
        cell_val = ws2.cell(conclusion_row, 2)
        fill_color = "D4EDDA" if int(n_tel) > 0 else "F8D7DA"
        cell_val.fill = PatternFill("solid", fgColor=fill_color)
        cell_val.font = Font(bold=True)
    except Exception:
        pass

    wb.save(filepath)
    return int(n_tel), int(n_email), total


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  TEST ENRICHISSEMENT — Pappers contacts_generiques")
    print(f"  Endpoint : {BASE_URL}")
    print(f"  Champ    : champs_supplementaires=contacts_generiques")
    print(f"  Budget   : {MAX_SIRETS} crédits max (1 crédit / appel)")
    print("=" * 65)

    # Chargement des SIRETs
    print("\n── Chargement des SIRETs ─────────────────────────────────────")
    siret_list = load_sirets()
    print(f"  Total SIRETs à tester : {len(siret_list)}")

    if not siret_list:
        print("[ERREUR] Aucun SIRET disponible.")
        sys.exit(1)

    # Appels API
    print(f"\n── Appels API ({len(siret_list)} × 1 crédit) ──────────────────────────")
    results = []
    n_tel = n_email = n_contacts_data = n_errors = 0

    for i, (siret, nom_source, fichier) in enumerate(siret_list, start=1):
        result = fetch_contacts(siret)

        if result["status"] == "RATE_LIMIT":
            print(f"  [{i:3d}] ⏳ Rate limit — pause 3s…")
            time.sleep(3)
            result = fetch_contacts(siret)   # 1 retry

        status_icon = "✅" if result["status"] == "OK" else "❌"
        has_tel   = bool(result.get("telephone"))
        has_email = bool(result.get("email"))
        has_data  = result.get("contacts_raw", "null") != "null"

        if result["status"] != "OK":
            n_errors += 1
        if has_tel:
            n_tel += 1
        if has_email:
            n_email += 1
        if has_data:
            n_contacts_data += 1

        # Log ligne par ligne
        flags = []
        if has_tel:   flags.append(f"📞 {result['telephone']}")
        if has_email: flags.append(f"✉️  {result['email']}")
        if has_data and not has_tel and not has_email:
            flags.append(f"⚠️ contacts≠null mais vide ({result['contacts_raw'][:40]})")
        if not has_data:
            flags.append("contacts_generiques=null")

        nom_pappers = result.get("nom_entreprise") or nom_source
        print(
            f"  [{i:3d}] {status_icon} {siret} | "
            f"{nom_pappers[:30]:<30} | "
            + (" | ".join(flags) if flags else "—")
        )

        results.append({
            "SIRET":                        siret,
            "Raison sociale (source)":      nom_source,
            "Raison sociale (Pappers)":     result.get("nom_entreprise") or "",
            "Source fichier":               fichier,
            "Statut Pappers":               result.get("statut_consolide") or result["status"],
            "contacts_generiques (brut)":   result.get("contacts_raw") or result["status"],
            "Nb contacts":                  result.get("nb_contacts") or 0,
            "Téléphone":                    result.get("telephone") or "",
            "Email":                        result.get("email") or "",
            "Résultat appel":               result["status"],
        })

        # Barre de progression rapide (tous les 10)
        if i % 10 == 0:
            pct_tel = n_tel / i * 100
            print(f"\n  ── {i}/{len(siret_list)} traités │ "
                  f"📞 {n_tel} tél ({pct_tel:.1f}%) │ "
                  f"✉️  {n_email} email │ "
                  f"data≠null: {n_contacts_data} │ "
                  f"erreurs: {n_errors} ──\n")

        time.sleep(DELAY)

    # Synthèse finale console
    total = len(siret_list)
    pct_tel   = n_tel / total * 100
    pct_email = n_email / total * 100

    print("\n" + "=" * 65)
    print("  RÉSULTATS FINAUX")
    print("=" * 65)
    print(f"  SIRETs testés              : {total}")
    print(f"  Appels réussis (HTTP 200)  : {total - n_errors}")
    print(f"  Erreurs                    : {n_errors}")
    print()
    print(f"  contacts_generiques ≠ null : {n_contacts_data} / {total}")
    print(f"  Avec TÉLÉPHONE             : {n_tel} / {total}  ({pct_tel:.1f}%)")
    print(f"  Avec EMAIL                 : {n_email} / {total}  ({pct_email:.1f}%)")
    print()
    if n_tel == 0 and n_contacts_data == 0:
        print("  ⛔ VERDICT : contacts_generiques toujours null")
        print("     → Pappers ne fournit PAS de téléphone via cette API")
        print("     → Le champ nécessite probablement un abonnement premium")
        print("       ou n'est tout simplement pas collecté par Pappers")
    elif n_tel == 0 and n_contacts_data > 0:
        print(f"  ⚠️  VERDICT : contacts_generiques parfois non-null ({n_contacts_data} cas)")
        print("     mais aucun téléphone/email extrait (structure inattendue)")
        print("     → Voir la colonne 'contacts_generiques (brut)' dans le fichier Excel")
    else:
        print(f"  ✅ VERDICT : Pappers fournit le téléphone pour {pct_tel:.1f}% des entreprises")
        print(f"     → Potentiellement utile pour l'enrichissement du pipeline")
    print("=" * 65)

    # Export
    n_tel_xlsx, n_email_xlsx, total_xlsx = export_excel(results, OUTPUT_FILE)
    print(f"\n[OK] Export : {OUTPUT_FILE}")
    print(f"     {total_xlsx} lignes │ 📞 {n_tel_xlsx} tél │ ✉️  {n_email_xlsx} email")


if __name__ == "__main__":
    main()
