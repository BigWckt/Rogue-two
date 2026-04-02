#!/usr/bin/env python3
"""
Scraper de prospection boulangeries — Skill & You
Source : API Annuaire des Entreprises (recherche-entreprises.api.gouv.fr)

Usage:
    python scraper_boulangeries.py
    python scraper_boulangeries.py --output ./resultats
"""

import argparse
import csv
import os
import sys
import time
from collections import Counter
from datetime import date

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

TODAY = date.today().strftime("%Y-%m-%d")

VILLES_CONFIG = [
    {"nom": "Lyon", "proprietaire": "Lina",
     "codes_commune": ["69381", "69382", "69383", "69384", "69385",
                        "69386", "69387", "69388", "69389"]},
    {"nom": "Drancy", "proprietaire": "Nabil", "codes_commune": ["93029"]},
    {"nom": "Le Blanc-Mesnil", "proprietaire": "Nabil", "codes_commune": ["93007"]},
    {"nom": "Pantin", "proprietaire": "Nabil", "codes_commune": ["93055"]},
    {"nom": "Bobigny", "proprietaire": "Nabil", "codes_commune": ["93008"]},
    {"nom": "Saint-Ouen-sur-Seine", "proprietaire": "Nabil", "codes_commune": ["93070"]},
    {"nom": "Montpellier", "proprietaire": "Camille", "codes_commune": ["34172"]},
    {"nom": "Boulogne-Billancourt", "proprietaire": "Aude", "codes_commune": ["92012"]},
    {"nom": "Nanterre", "proprietaire": "Aude", "codes_commune": ["92050"]},
    {"nom": "Rueil-Malmaison", "proprietaire": "Aude", "codes_commune": ["92063"]},
    {"nom": "Lille", "proprietaire": "Samya", "codes_commune": ["59350"]},
]

NAF_BOULANGERIE = ["10.71C", "10.71D", "47.24Z"]

API_BASE = "https://recherche-entreprises.api.gouv.fr/search"
API_DELAY = 1.5
API_RETRIES = 3
PER_PAGE = 25

# ── Fonctions ─────────────────────────────────────────────────────────────────


def api_get(params, retries=API_RETRIES):
    for attempt in range(retries):
        try:
            r = requests.get(API_BASE, params=params, timeout=20)
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429:
                wait = int(r.headers.get("retry-after", 4)) + 1
                print(f"  Rate limited, attente {wait}s...")
                time.sleep(wait)
                continue
            if r.status_code >= 500:
                print(f"  Erreur serveur {r.status_code}, retry {attempt+1}/{retries}")
                time.sleep(2 ** (attempt + 1))
                continue
            print(f"  Erreur API {r.status_code}: {r.text[:200]}")
            return None
        except requests.exceptions.RequestException as e:
            print(f"  Erreur reseau: {e}, retry {attempt+1}/{retries}")
            time.sleep(2 ** (attempt + 1))
    return None


def fetch_etablissements(code_commune, naf_code):
    etablissements = []
    page = 1
    while True:
        params = {
            "activite_principale": naf_code,
            "code_commune": code_commune,
            "etat_administratif": "A",
            "page": page,
            "per_page": PER_PAGE,
        }
        time.sleep(API_DELAY)
        data = api_get(params)
        if not data or not data.get("results"):
            break

        for entreprise in data["results"]:
            siren = entreprise.get("siren", "")
            nom_entreprise = entreprise.get("nom_complet", "")

            for etab in entreprise.get("matching_etablissements", []):
                if etab.get("etat_administratif") != "A":
                    continue
                if etab.get("commune") != code_commune:
                    continue
                enseignes = etab.get("liste_enseignes") or []
                nom = enseignes[0] if enseignes else (etab.get("nom_commercial") or nom_entreprise)
                etablissements.append({
                    "Nom établissement": nom.strip().title() if nom else "",
                    "Adresse": etab.get("adresse", ""),
                    "Code Postal": etab.get("code_postal", ""),
                    "Ville": etab.get("libelle_commune", ""),
                    "SIRET": etab.get("siret", ""),
                    "SIREN": siren,
                    "NAF": etab.get("activite_principale", naf_code),
                })

            if not entreprise.get("matching_etablissements"):
                siege = entreprise.get("siege", {})
                if siege.get("commune") == code_commune and siege.get("etat_administratif") == "A":
                    enseignes = siege.get("liste_enseignes") or []
                    nom = enseignes[0] if enseignes else (siege.get("nom_commercial") or nom_entreprise)
                    etablissements.append({
                        "Nom établissement": nom.strip().title() if nom else "",
                        "Adresse": siege.get("adresse", ""),
                        "Code Postal": siege.get("code_postal", ""),
                        "Ville": siege.get("libelle_commune", ""),
                        "SIRET": siege.get("siret", ""),
                        "SIREN": siren,
                        "NAF": siege.get("activite_principale", naf_code),
                    })

        total_pages = data.get("total_pages", 1)
        if page >= total_pages:
            break
        page += 1
    return etablissements


def scrape_ville(ville_config):
    nom_ville = ville_config["nom"]
    codes_commune = ville_config["codes_commune"]
    proprietaire = ville_config["proprietaire"]

    print(f"\nScraping {nom_ville} ({len(codes_commune)} commune(s))...")

    all_etabs = []
    for code in codes_commune:
        for naf in NAF_BOULANGERIE:
            etabs = fetch_etablissements(code, naf)
            all_etabs.extend(etabs)
            if etabs:
                print(f"  Commune {code} / NAF {naf}: {len(etabs)} etablissements")

    seen_sirets = set()
    unique = []
    for e in all_etabs:
        siret = e.get("SIRET", "")
        if siret and siret in seen_sirets:
            continue
        if siret:
            seen_sirets.add(siret)
        key = (e["Nom établissement"].lower(), e["Code Postal"])
        if not siret and key in seen_sirets:
            continue
        seen_sirets.add(key)
        unique.append(e)

    for e in unique:
        e["Type d'établissement"] = "Boulangerie"
        e["Priorité"] = 3
        e["Propriétaire"] = proprietaire
        e["Numéro de téléphone"] = ""

    print(f"  Total {nom_ville}: {len(unique)} etablissements uniques")
    return unique


def save_backup_csv(data, filepath):
    if not data:
        return
    keys = ["Nom établissement", "Adresse", "Numéro de téléphone",
            "Type d'établissement", "Priorité", "Propriétaire", "SIREN"]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)
    print(f"  Backup CSV: {filepath}")


def export_excel(all_data, output_path):
    columns = [
        "Nom établissement",
        "Adresse",
        "Numéro de téléphone",
        "Type d'établissement",
        "Priorité",
        "Propriétaire",
        "SIREN",
    ]

    df = pd.DataFrame(all_data, columns=columns)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Prospection")
        ws = writer.sheets["Prospection"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        for row in range(2, ws.max_row + 1):
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(vertical="center")
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")

    print(f"\nExport Excel: {output_path}")
    print(f"  {len(df)} lignes exportees")


def main():
    parser = argparse.ArgumentParser(description="Scraper boulangeries — Skill & You")
    parser.add_argument("--output", default=".", help="Repertoire de sortie")
    args = parser.parse_args()

    output_dir = args.output
    os.makedirs(output_dir, exist_ok=True)

    xlsx_path = os.path.join(output_dir, f"boulangeries_prospection_{TODAY}.xlsx")
    csv_backup_path = os.path.join(output_dir, f"boulangeries_backup_{TODAY}.csv")

    print("=" * 60)
    print(f"Scraper Boulangeries - Prospection Skill & You")
    print(f"  Date: {TODAY}")
    print(f"  Villes: {len(VILLES_CONFIG)}")
    print(f"  Codes NAF: {', '.join(NAF_BOULANGERIE)}")
    print("=" * 60)

    all_results = []
    for ville in VILLES_CONFIG:
        try:
            results = scrape_ville(ville)
            all_results.extend(results)
            save_backup_csv(all_results, csv_backup_path)
        except Exception as e:
            print(f"  Erreur sur {ville['nom']}: {e}")
            if all_results:
                save_backup_csv(all_results, csv_backup_path)

    if not all_results:
        print("\nAucun resultat trouve.")
        sys.exit(1)

    seen = set()
    deduped = []
    for r in all_results:
        siret = r.get("SIRET", "")
        if siret and siret in seen:
            continue
        if siret:
            seen.add(siret)
        deduped.append(r)

    print(f"\nSynthese finale:")
    print(f"  Total brut: {len(all_results)}")
    print(f"  Apres dedup: {len(deduped)}")

    villes_count = Counter()
    propri_count = Counter()
    for r in deduped:
        villes_count[r.get("Ville", "?")] += 1
        propri_count[r.get("Propriétaire", "?")] += 1

    print("\n  Par ville:")
    for v, c in villes_count.most_common():
        print(f"    {v}: {c}")
    print("\n  Par proprietaire:")
    for p, c in propri_count.most_common():
        print(f"    {p}: {c}")

    export_excel(deduped, xlsx_path)
    save_backup_csv(deduped, csv_backup_path)
    print(f"\nTermine ! Fichier: {xlsx_path}")


if __name__ == "__main__":
    main()
