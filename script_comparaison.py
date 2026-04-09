#!/usr/bin/env python3
"""
script_comparaison.py — Croisement PJ enrichi x LBA/LBB
========================================================
Croise les résultats Pages Jaunes enrichis (avec SIRET) et les résultats
LBA/LBB sur la base du SIRET, attribue un niveau de priorité (Haute /
Moyenne / Basse), et produit le fichier Excel final multi-onglets prêt
pour import HubSpot.

Usage :
  python script_comparaison.py --pj pj_enrichi.xlsx --lba lba_lbb.xlsx
  python script_comparaison.py --lba lba_lbb.xlsx  # sans PJ
"""

import argparse
import os
import sys
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Colonnes de sortie (noms HubSpot) ────────────────────────────────────────

OUTPUT_COLUMNS = [
    "Nom de l'entreprise",
    "SIRET",
    "Code NAF",
    "Adresse",
    "Ville",
    "Code Postal",
    "Téléphone",
    "Site web",
    "Priorité",
    "Source",
    "Score LBB",
    "Offres actives",
    "Date de collecte",
]

# ── Charte Skill & You ──────────────────────────────────────────────────────

HEADER_BG = "1558EE"
HEADER_FG = "FFFFFF"

PRIORITY_ORDER = {"Haute": 0, "Moyenne": 1, "Basse": 2}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_siret(raw) -> str:
    if raw is None or str(raw).strip() in ("", "nan", "None"):
        return ""
    try:
        return str(int(float(str(raw).strip()))).zfill(14)
    except (ValueError, OverflowError):
        return str(raw).strip()


def load_lba_lbb(filepath: str) -> dict[str, dict]:
    """Charge le fichier LBA/LBB et retourne un dict SIRET → row."""
    try:
        df = pd.read_excel(filepath, sheet_name="Consolidé", dtype=str).fillna("")
    except (ValueError, KeyError):
        df = pd.read_excel(filepath, dtype=str).fillna("")
    by_siret: dict[str, dict] = {}
    for _, row in df.iterrows():
        siret = normalize_siret(row.get("SIRET", ""))
        if not siret:
            continue
        rec = row.to_dict()
        rec["SIRET"] = siret
        # Si doublon SIRET, garder celui avec source LBA (prioritaire)
        if siret in by_siret:
            existing_src = by_siret[siret].get("Source", "")
            if "LBA" not in existing_src and "LBA" in rec.get("Source", ""):
                by_siret[siret] = rec
        else:
            by_siret[siret] = rec
    return by_siret


def load_pj_enrichi(filepath: str) -> dict[str, dict]:
    """Charge le fichier PJ enrichi (onglet Entreprises) et retourne un dict SIRET → row."""
    try:
        df = pd.read_excel(filepath, sheet_name="Entreprises", dtype=str).fillna("")
    except (ValueError, KeyError):
        df = pd.read_excel(filepath, sheet_name=0, dtype=str).fillna("")
    by_siret: dict[str, dict] = {}
    for _, row in df.iterrows():
        siret = normalize_siret(row.get("SIRET", ""))
        if not siret:
            continue
        rec = row.to_dict()
        rec["SIRET"] = siret
        if siret not in by_siret:
            by_siret[siret] = rec
    return by_siret


# ── Croisement ────────────────────────────────────────────────────────────────

def merge_and_score(pj_data: dict[str, dict],
                    lba_data: dict[str, dict]) -> tuple[list[dict], dict]:
    """
    Croise PJ et LBA/LBB par SIRET, attribue la priorité.
    Retourne (leads, stats).
    """
    today = date.today().isoformat()
    leads: list[dict] = []
    stats = {
        "pj_count": len(pj_data),
        "lba_count": len(lba_data),
        "matches": 0,
        "haute": 0,
        "moyenne": 0,
        "basse": 0,
    }

    all_sirets = set(pj_data.keys()) | set(lba_data.keys())

    for siret in all_sirets:
        pj = pj_data.get(siret)
        lba = lba_data.get(siret)

        lead: dict = {col: "" for col in OUTPUT_COLUMNS}
        lead["SIRET"] = siret
        lead["Date de collecte"] = today

        if pj and lba:
            # Match : fusionner — PJ pour nom/adresse/tel, LBA/LBB pour score/offres
            stats["matches"] += 1
            lead["Nom de l'entreprise"] = pj.get("Nom de l'entreprise") or lba.get("Nom de l'entreprise", "")
            lead["Code NAF"] = pj.get("Code NAF") or lba.get("Code NAF", "")
            lead["Adresse"] = pj.get("Adresse") or lba.get("Adresse", "")
            lead["Ville"] = pj.get("Ville") or lba.get("Ville", "")
            lead["Code Postal"] = pj.get("Code Postal") or lba.get("Code Postal", "")
            lead["Téléphone"] = pj.get("Téléphone", "")
            lead["Site web"] = pj.get("Site web", "")
            lead["Score LBB"] = lba.get("Score LBB", "")
            lead["Offres actives"] = lba.get("Offres actives", "")

            lba_source = lba.get("Source", "")
            if "LBA" in lba_source:
                lead["Priorité"] = "Haute"
                lead["Source"] = "PJ + LBA" if "LBB" not in lba_source else "PJ + LBA + LBB"
            else:
                lead["Priorité"] = "Moyenne"
                lead["Source"] = "PJ + LBB"

        elif lba:
            # Uniquement LBA/LBB — pas de correspondance PJ
            lead["Nom de l'entreprise"] = lba.get("Nom de l'entreprise", "")
            lead["Code NAF"] = lba.get("Code NAF", "")
            lead["Adresse"] = lba.get("Adresse", "")
            lead["Ville"] = lba.get("Ville", "")
            lead["Code Postal"] = lba.get("Code Postal", "")
            lead["Score LBB"] = lba.get("Score LBB", "")
            lead["Offres actives"] = lba.get("Offres actives", "")

            lba_source = lba.get("Source", "")
            if "LBA" in lba_source:
                lead["Priorité"] = "Haute"
                lead["Source"] = lba_source
            else:
                lead["Priorité"] = "Moyenne"
                lead["Source"] = lba_source

        elif pj:
            # Uniquement PJ — aucun signal recrutement
            lead["Nom de l'entreprise"] = pj.get("Nom de l'entreprise", "")
            lead["Code NAF"] = pj.get("Code NAF", "")
            lead["Adresse"] = pj.get("Adresse", "")
            lead["Ville"] = pj.get("Ville", "")
            lead["Code Postal"] = pj.get("Code Postal", "")
            lead["Téléphone"] = pj.get("Téléphone", "")
            lead["Site web"] = pj.get("Site web", "")
            lead["Priorité"] = "Basse"
            lead["Source"] = "Pages Jaunes"

        # Compteurs
        if lead["Priorité"] == "Haute":
            stats["haute"] += 1
        elif lead["Priorité"] == "Moyenne":
            stats["moyenne"] += 1
        elif lead["Priorité"] == "Basse":
            stats["basse"] += 1

        leads.append(lead)

    # Tri : Haute → Moyenne → Basse
    leads.sort(key=lambda r: PRIORITY_ORDER.get(r.get("Priorité", ""), 9))

    return leads, stats


# ── Export Excel ──────────────────────────────────────────────────────────────

def style_sheet(ws):
    """Applique la mise en forme Skill & You à un onglet."""
    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    hdr_font = Font(bold=True, color=HEADER_FG)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-dimensionner les colonnes
    for col_idx in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(1, col_idx).value or ""))
        for row_idx in range(2, min(ws.max_row + 1, 50)):  # sample 50 rows
            val = str(ws.cell(row_idx, col_idx).value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    # SIRET en texte
    siret_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(1, col_idx).value == "SIRET":
            siret_col = col_idx
            break
    if siret_col:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, siret_col).number_format = "@"

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def export_excel(leads: list[dict], output_path: str):
    """Exporte le fichier Excel multi-onglets."""
    haute = [r for r in leads if r["Priorité"] == "Haute"]
    moyenne = [r for r in leads if r["Priorité"] == "Moyenne"]
    basse = [r for r in leads if r["Priorité"] == "Basse"]

    sheets = {
        "Priorité Haute": haute,
        "Priorité Moyenne": moyenne,
        "Priorité Basse": basse,
        "Tous les leads": leads,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows in sheets.items():
            df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
            df["SIRET"] = df["SIRET"].astype(str)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Styling
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])
    wb.save(output_path)


# ── Synthèse console ─────────────────────────────────────────────────────────

def print_synthese(ville: str, leads: list[dict], stats: dict,
                   output_path: str):
    n_total = len(leads)
    print()
    print("══════════════════════════════════════════════════")
    print(f"  CROISEMENT — {ville}")
    if stats["pj_count"]:
        print(f"  PJ enrichi  : {stats['pj_count']} entreprises (avec SIRET)")
    print(f"  LBA/LBB     : {stats['lba_count']} entreprises")
    print()
    if stats["pj_count"]:
        print(f"  Correspondances PJ ↔ LBA/LBB (même SIRET) : {stats['matches']}")
    print("  ──────────────────────────────────────────────")
    print(f"  🔴 Priorité Haute   : {stats['haute']} entreprises")
    print(f"  🟡 Priorité Moyenne : {stats['moyenne']} entreprises")
    print(f"  🟢 Priorité Basse   : {stats['basse']} entreprises")
    print("  ──────────────────────────────────────────────")
    print(f"  Total leads qualifiés : {n_total} entreprises")
    print()
    print(f"  📁 Fichier : {os.path.basename(output_path)}")
    print("══════════════════════════════════════════════════")


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Croisement PJ enrichi × LBA/LBB — scoring et export HubSpot",
    )
    parser.add_argument("--pj", type=str, default=None,
                        help="Fichier PJ enrichi (.xlsx)")
    parser.add_argument("--lba", type=str, default=None,
                        help="Fichier LBA/LBB (.xlsx)")
    parser.add_argument("--output", type=str, default=".",
                        help="Répertoire de sortie")
    return parser.parse_args()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()

    if not args.pj and not args.lba:
        print("❌ Spécifiez au moins --pj ou --lba")
        sys.exit(1)

    os.makedirs(args.output, exist_ok=True)

    # Charger les données
    pj_data: dict[str, dict] = {}
    lba_data: dict[str, dict] = {}

    if args.pj:
        if not os.path.exists(args.pj):
            print(f"⚠️  Fichier PJ introuvable : {args.pj} — continué sans PJ")
        else:
            print(f"📂 PJ enrichi : {args.pj}")
            pj_data = load_pj_enrichi(args.pj)
            print(f"  {len(pj_data)} entreprises avec SIRET")

    if args.lba:
        if not os.path.exists(args.lba):
            print(f"⚠️  Fichier LBA/LBB introuvable : {args.lba} — continué sans LBA/LBB")
        else:
            print(f"📂 LBA/LBB    : {args.lba}")
            lba_data = load_lba_lbb(args.lba)
            print(f"  {len(lba_data)} entreprises avec SIRET")

    if not pj_data and not lba_data:
        print("❌ Aucune donnée chargée")
        sys.exit(1)

    # Déduire la ville depuis le nom de fichier
    ville = "Inconnu"
    ref_file = args.lba or args.pj or ""
    base = os.path.basename(ref_file).lower()
    for prefix in ["lba_lbb_results_", "pj_results_"]:
        if prefix in base:
            parts = base.replace(prefix, "").split("_")
            if parts:
                ville = parts[0].title()
                break

    print(f"\n── Croisement par SIRET ─────────────────────────")

    # Croisement
    leads, stats = merge_and_score(pj_data, lba_data)

    if not leads:
        print("⚠️  Aucun lead généré")
        sys.exit(0)

    # Export
    today_str = date.today().strftime("%Y%m%d")
    ville_slug = ville.lower().replace(" ", "_").replace("-", "_")
    output_path = os.path.join(args.output, f"leads_qualifies_{ville_slug}_{today_str}.xlsx")

    print(f"\n── Export ──────────────────────────────────────")
    export_excel(leads, output_path)
    print(f"  ✅ {output_path}")

    # Synthèse
    print_synthese(ville, leads, stats, output_path)


if __name__ == "__main__":
    main()
