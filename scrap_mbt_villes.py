"""
Scrap MBT multi-villes — LBA + LBB
============================================================
CAP Boulanger / CAP Pâtissier / TP CV / CAP EPC
LBA : ROME D1101, D1102, D1103, D1104, D1106, G1603
LBB : NAF  10.71A/B/C/D (fallback ROME si non supporté)

10 villes (paramètres managers) :
  Paris 75001 — 50 km   |  Toulouse 31000 — 30 km
  Lille 59000 — 30 km   |  Lyon 69001     — 30 km
  Marseille 13001 — 30km|  Montpellier 34000 — 30 km
  Nice 06000 — 30 km    |  Bordeaux 33000 — 30 km
  Perpignan 66000 — 30km|  Toulon 83000   — 30 km

Règles :
  - 300 lignes max par ville après dédup SIRET local
  - Tri prioritaire : avec téléphone en premier
  - Dédup globale par SIRET (chaque entreprise dans une seule ville)
  - SIRET normalisé : str(int(float(siret))).zfill(14)
  - Export : 1 onglet par ville + onglet "Synthèse"
"""

import re as _re
import sys
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Villes managers ───────────────────────────────────────────────────────────
# (nom_affichage, code_postal_ref, latitude, longitude, rayon_km)

CITIES = [
    ("Paris",       "75001", 48.8603,  2.3477,  50),
    ("Toulouse",    "31000", 43.6047,  1.4442,  30),
    ("Lille",       "59000", 50.6292,  3.0573,  30),
    ("Lyon",        "69001", 45.7640,  4.8357,  30),
    ("Marseille",   "13001", 43.2965,  5.3698,  30),
    ("Montpellier", "34000", 43.6119,  3.8772,  30),
    ("Nice",        "06000", 43.7102,  7.2620,  30),
    ("Bordeaux",    "33000", 44.8378, -0.5792,  30),
    ("Perpignan",   "66000", 42.6887,  2.8948,  30),
    ("Toulon",      "83000", 43.1242,  5.9280,  30),
]

MAX_PER_CITY   = 300   # lignes max par ville dans l'export final
LBB_CITY_CAP   = 900   # SIRETs LBB collectés par ville (marge avant dédup globale)

# ── Codes MBT ─────────────────────────────────────────────────────────────────

ROME_CODES         = ["D1101", "D1102", "D1103", "D1104", "D1106", "G1603"]
NAF_CODES_LBB      = ["10.71A", "10.71B", "10.71C", "10.71D"]

# ── Authentification ──────────────────────────────────────────────────────────

LBB_ENDPOINT  = "https://api.francetravail.io/partenaire/labonneboite/v2/recherche"
LBB_OAUTH_URL = (
    "https://authentification-partenaire.francetravail.io"
    "/connexion/oauth2/access_token?realm=/partenaire"
)
LBB_OAUTH_URL_LEGACY = (
    "https://entreprise.francetravail.fr"
    "/connexion/oauth2/access_token?realm=/partenaire"
)
LBB_CLIENT_ID     = "PAR_claudecode_dbfd12ec4f6fe1174e46c36b762d98130ae05b4c33d069c1a5bebebe8573f33a"
LBB_CLIENT_SECRET = "a866591ed795a34b62c4d379e7f0cff3e393a59da7fbf3dfb04f101b90de2dd3"
LBB_SCOPE_CANDIDATES = [
    "search office api_labonneboitev2",
    f"search office api_labonneboitev2 application_{LBB_CLIENT_ID}",
]

LBA_ENDPOINT = "https://api.apprentissage.beta.gouv.fr/api/job/v1/search"
LBA_TOKEN = (
    "eyJhbGciOiJIUzI1NiJ9."
    "eyJfaWQiOiI2OGM4MzJkZDgxZGY5MmFiYTc2MDNhNzAiLCJhcGlfa2V5IjoieVFaYkpiZElCN1VydDNvb2"
    "w3aTRqN0lSRUhSQ25KSk5ya0djclpxZ1E0bz0iLCJvcmdhbmlzYXRpb24iOiJTa2lsbGFuZHlvdSIsImVt"
    "YWlsIjoiam9mZnJleS5sYWRtaXJhdWx0QHNraWxsYW5keW91LmNvbSIsImlzcyI6ImFwaSIsImlhdCI6MTc"
    "3MzMyMzg1NSwiZXhwIjoxNzg5NDg2Njk2fQ."
    "5P4D4gmCTezPIvycQEtKnPZBjPvcx8BHG1bF8r_Z4ts"
)

# ── Pappers (estimation uniquement) ───────────────────────────────────────────

PAPPERS_CREDITS_PER_SIRET = 4
PAPPERS_MONTHLY_QUOTA     = 2_000

# ── Sortie ────────────────────────────────────────────────────────────────────

OUTPUT_FILE   = "mbt_villes.xlsx"
EXCEL_COLUMNS = [
    "Source", "Type", "SIRET", "Raison sociale",
    "Adresse", "Code postal", "Ville",
    "Téléphone", "Email",
    "Code NAF", "Code ROME", "Distance (km)",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_siret(raw) -> str:
    if raw is None or str(raw).strip() in ("", "nan", "None"):
        return ""
    try:
        return str(int(float(str(raw).strip()))).zfill(14)
    except (ValueError, OverflowError):
        return str(raw).strip()


def http_get(url, params=None, headers=None, timeout=20):
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        if resp.status_code == 429:
            print("  [WARN] 429 — pause 10s…")
            time.sleep(10)
            resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        if resp.status_code in (401, 403):
            print(f"  [ERROR] HTTP {resp.status_code} — {url}")
            return None
        resp.raise_for_status()
        if "html" in resp.headers.get("content-type", ""):
            print(f"  [ERROR] Réponse HTML — {url}")
            return None
        return resp.json()
    except requests.exceptions.Timeout:
        print(f"  [ERROR] Timeout — {url}")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"  [ERROR] Connexion : {e}")
        return None
    except (requests.exceptions.HTTPError, ValueError) as e:
        print(f"  [ERROR] {e}")
        return None


def _parse_address(full_address: str) -> tuple[str, str, str]:
    if not full_address:
        return "", "", ""
    m = _re.search(r'(\d{5})\s+(.+)$', full_address.strip())
    if m:
        return (
            full_address[:m.start()].strip().rstrip(',').strip(),
            m.group(1),
            m.group(2).strip(),
        )
    return full_address, "", ""


# ── OAuth LBB ─────────────────────────────────────────────────────────────────

_lbb_token_cache: dict = {}


def _get_lbb_token() -> str | None:
    now = time.time()
    if _lbb_token_cache.get("token") and _lbb_token_cache.get("expires_at", 0) > now + 30:
        return _lbb_token_cache["token"]

    print("  [LBB] Obtention token OAuth2…")
    for scope in LBB_SCOPE_CANDIDATES:
        for label, oauth_url in [("ProxyProConnect", LBB_OAUTH_URL), ("Legacy", LBB_OAUTH_URL_LEGACY)]:
            try:
                resp = requests.post(oauth_url, data={
                    "grant_type": "client_credentials",
                    "client_id": LBB_CLIENT_ID,
                    "client_secret": LBB_CLIENT_SECRET,
                    "scope": scope,
                }, timeout=15)
                if resp.status_code in (400, 401):
                    continue
                resp.raise_for_status()
                token      = resp.json().get("access_token")
                expires_in = int(resp.json().get("expires_in", 1800))
                probe = requests.get(LBB_ENDPOINT, params={
                    "latitude": 48.8603, "longitude": 2.3477,
                    "distance": 1, "rome": "D1102", "page_size": 1,
                }, headers={"Authorization": f"Bearer {token}"}, timeout=10)
                if probe.status_code == 200:
                    print(f"  [LBB] ✅ Token valide ({label})")
                    _lbb_token_cache["token"]      = token
                    _lbb_token_cache["expires_at"] = now + expires_in
                    return token
            except Exception:
                continue
    print("  [LBB][ERROR] Aucun token LBB valide obtenu")
    return None


# ── LBB fetch (par ville) ────────────────────────────────────────────────────

def fetch_lbb_city(token: str, city_name: str, lat: float, lon: float,
                   dist_km: int) -> list[dict]:
    """
    Collecte les entreprises LBB pour une ville.
    Tente d'abord filtre NAF ; fallback ROME si vide.
    S'arrête dès que LBB_CITY_CAP SIRETs uniques sont collectés.
    """
    headers = {"Authorization": f"Bearer {token}"}

    def _collect(param_name: str, param_values: list[str]) -> list[dict]:
        results: list[dict]  = []
        seen_sirets: set[str] = set()

        for val in param_values:
            if len(seen_sirets) >= LBB_CITY_CAP:
                break
            page = 1
            page_size = 100
            while len(seen_sirets) < LBB_CITY_CAP:
                data = http_get(LBB_ENDPOINT, params={
                    "latitude": lat, "longitude": lon,
                    "distance": dist_km,
                    param_name: val,
                    "page": page,
                    "page_size": page_size,
                }, headers=headers)
                if data is None:
                    break
                items = data.get("items") or []
                if not items:
                    break
                for c in items:
                    siret = normalize_siret(c.get("siret") or "")
                    if not siret or siret in seen_sirets:
                        continue
                    seen_sirets.add(siret)
                    results.append({
                        "Source": "LBB",
                        "Type": "Entreprise cible",
                        "SIRET": siret,
                        "Raison sociale": c.get("company_name") or c.get("office_name") or "",
                        "Adresse": "",
                        "Code postal": str(c.get("postcode") or ""),
                        "Ville": c.get("city") or city_name,
                        "Téléphone": "",
                        "Email": "",
                        "Code NAF": c.get("naf") or "",
                        "Code ROME": val if param_name == "rome" else "",
                        "Distance (km)": "",
                    })
                if len(items) < page_size:
                    break
                page += 1
                time.sleep(0.2)
        return results

    # Tenter NAF d'abord
    naf_results = _collect("naf", NAF_CODES_LBB)
    if naf_results:
        return naf_results

    # Fallback ROME
    return _collect("rome", ROME_CODES)


# ── LBA fetch (par ville) ─────────────────────────────────────────────────────

def fetch_lba_city(city_name: str, lat: float, lon: float,
                   dist_km: int) -> list[dict]:
    """Collecte toutes les offres + recruteurs LBA pour une ville."""
    headers = {"Authorization": f"Bearer {LBA_TOKEN}"}
    seen: dict[str, dict] = {}

    for rome in ROME_CODES:
        data = http_get(LBA_ENDPOINT, params={
            "romes": rome, "latitude": lat,
            "longitude": lon, "radius": dist_km,
        }, headers=headers)
        if data is None:
            continue

        jobs       = data.get("jobs")       or []
        recruiters = data.get("recruiters") or []

        # Recruteurs en priorité basse
        for item in recruiters:
            company = _parse_lba_item(item, rome, city_name, "Entreprise cible")
            if company and company["SIRET"] not in seen:
                seen[company["SIRET"]] = company

        # Offres actives en priorité haute
        for item in jobs:
            company = _parse_lba_item(item, rome, city_name, "Offre active")
            if company:
                seen[company["SIRET"]] = company  # écrase recruteur

        time.sleep(0.2)

    return list(seen.values())


def _parse_lba_item(item: dict, rome: str, city_name: str, type_: str) -> dict | None:
    wp    = item.get("workplace") or {}
    apply = item.get("apply")     or {}
    naf   = (wp.get("domain") or {}).get("naf") or {}

    siret = normalize_siret(wp.get("siret") or "")
    if not siret:
        return None

    offer = item.get("offer") or {}
    rome_final = (offer.get("rome_codes") or [rome])[0]

    rue, cp, ville = _parse_address(
        (wp.get("location") or {}).get("address") or ""
    )

    return {
        "Source": "LBA",
        "Type": type_,
        "SIRET": siret,
        "Raison sociale": wp.get("name") or wp.get("legal_name") or wp.get("brand") or "",
        "Adresse": rue,
        "Code postal": cp,
        "Ville": ville or city_name,
        "Téléphone": apply.get("phone") or "",
        "Email": "",
        "Code NAF": naf.get("code") or "",
        "Code ROME": rome_final,
        "Distance (km)": "",
    }


# ── Fusion locale (LBB + LBA) ─────────────────────────────────────────────────

def merge_local(lbb_rows: list[dict], lba_rows: list[dict]) -> dict[str, dict]:
    """
    Fusionne LBB + LBA par SIRET pour une ville.
    LBA prioritaire (téléphone, type Offre active).
    Retourne un dict SIRET → row.
    """
    by_siret: dict[str, dict] = {}

    for row in lbb_rows:
        s = row["SIRET"]
        if s:
            by_siret[s] = row.copy()

    for row in lba_rows:
        s = row["SIRET"]
        if not s:
            continue
        if s not in by_siret:
            by_siret[s] = row.copy()
        else:
            existing = by_siret[s]
            sources = set(existing["Source"].split(" + ")) | {"LBA"}
            existing["Source"] = " + ".join(sorted(sources))
            if row["Type"] == "Offre active":
                existing["Type"] = "Offre active"
            for col in EXCEL_COLUMNS:
                if not str(existing.get(col, "")).strip() and str(row.get(col, "")).strip():
                    existing[col] = row[col]

    return by_siret


# ── Export Excel ──────────────────────────────────────────────────────────────

def _style_sheet(ws, rows_with_phone: int, total_rows: int) -> None:
    """Applique la mise en forme à un onglet de données."""
    hdr_fill  = PatternFill("solid", fgColor="1A3A6B")
    ok_fill   = PatternFill("solid", fgColor="D4EDDA")
    ko_fill   = PatternFill("solid", fgColor="FFF3CD")
    alt_fill  = PatternFill("solid", fgColor="EBF3FF")
    wht_fill  = PatternFill("solid", fgColor="FFFFFF")

    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    tel_col = EXCEL_COLUMNS.index("Téléphone") + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        base = wht_fill if row_idx % 2 == 0 else alt_fill
        for cell in row:
            cell.fill      = base
            cell.alignment = Alignment(vertical="center")
        tel_cell = ws.cell(row_idx, tel_col)
        if str(tel_cell.value or "").strip():
            tel_cell.fill = ok_fill
            tel_cell.font = Font(bold=True, color="1A7A1A")
        else:
            tel_cell.fill = ko_fill

    col_widths = [14, 16, 16, 38, 35, 12, 22, 16, 28, 12, 12, 14]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def export_excel(city_data: dict[str, list[dict]], filepath: str) -> None:
    """
    Exporte un onglet par ville + un onglet Synthèse.
    city_data : {city_name: [row, ...]}
    """
    # Aplatir toutes les lignes pour la synthèse
    all_rows: list[dict] = []
    for rows in city_data.values():
        all_rows.extend(rows)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Onglet par ville
        for city_name, rows in city_data.items():
            if not rows:
                continue
            df_city = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
            sheet_name = city_name[:31]  # limite openpyxl
            df_city.to_excel(writer, sheet_name=sheet_name, index=False)

        # Onglet Synthèse
        df_all = pd.DataFrame(all_rows, columns=EXCEL_COLUMNS)
        df_all.to_excel(writer, sheet_name="Synthèse", index=False)

    wb = load_workbook(filepath)

    for city_name, rows in city_data.items():
        if not rows:
            continue
        ws = wb[city_name[:31]]
        n_tel = sum(1 for r in rows if str(r.get("Téléphone", "")).strip())
        _style_sheet(ws, n_tel, len(rows))

    ws_synth = wb["Synthèse"]
    n_tel_all = sum(1 for r in all_rows if str(r.get("Téléphone", "")).strip())
    _style_sheet(ws_synth, n_tel_all, len(all_rows))

    # Réorganiser : Synthèse en premier
    sheet_names = wb.sheetnames
    if "Synthèse" in sheet_names:
        idx = sheet_names.index("Synthèse")
        wb.move_sheet("Synthèse", offset=-idx)

    wb.save(filepath)


# ── Rapport bilan ─────────────────────────────────────────────────────────────

def print_rapport(city_data: dict[str, list[dict]]) -> None:
    all_rows = [r for rows in city_data.values() for r in rows]
    total    = len(all_rows)
    n_tel    = sum(1 for r in all_rows if str(r.get("Téléphone", "")).strip())
    n_no_tel = total - n_tel

    credits_needed      = n_no_tel * PAPPERS_CREDITS_PER_SIRET
    max_enrich_month    = PAPPERS_MONTHLY_QUOTA // PAPPERS_CREDITS_PER_SIRET
    enrichissables      = min(n_no_tel, max_enrich_month)
    tel_attendus        = int(enrichissables * 0.20)

    print("\n" + "=" * 65)
    print("  BILAN DU SCRAP MBT — 10 VILLES")
    print("=" * 65)
    print(f"  {'Ville':<14} {'Total':>6}  {'Avec tél':>8}  {'Sans tél':>8}")
    print("  " + "─" * 42)
    for city_name, rows in city_data.items():
        t   = len(rows)
        tel = sum(1 for r in rows if str(r.get("Téléphone", "")).strip())
        print(f"  {city_name:<14} {t:>6}  {tel:>8}  {t-tel:>8}")
    print("  " + "─" * 42)
    print(f"  {'TOTAL':<14} {total:>6}  {n_tel:>8}  {n_no_tel:>8}")
    print()
    print(f"  AVEC téléphone   : {n_tel}  ({n_tel/max(total,1)*100:.1f}%)")
    print(f"  SANS téléphone   : {n_no_tel}  ({n_no_tel/max(total,1)*100:.1f}%)")

    print()
    print("─" * 65)
    print("  ESTIMATION ENRICHISSEMENT PAPPERS")
    print("─" * 65)
    print(f"  Lignes sans téléphone          : {n_no_tel}")
    print(f"  Coût par SIRET                 : {PAPPERS_CREDITS_PER_SIRET} crédits")
    print(f"  Coût total si 100% enrichi     : {credits_needed:,} crédits")
    print()
    print(f"  Quota mensuel                  : {PAPPERS_MONTHLY_QUOTA:,} crédits")
    print(f"  Max enrichissements / mois     : {max_enrich_month} SIRETs")
    print()
    if n_no_tel <= max_enrich_month:
        print(f"  ✅ Budget suffisant ({n_no_tel} ≤ {max_enrich_month})")
        print(f"     Crédits consommés  : {credits_needed:,}")
        print(f"     Crédits restants   : {PAPPERS_MONTHLY_QUOTA - credits_needed:,}")
        print("  → RECOMMANDATION : enrichissement possible en 1 seul batch")
    else:
        mois = -(-n_no_tel // max_enrich_month)   # division plafond
        print(f"  ⚠️  Budget insuffisant ({n_no_tel} > {max_enrich_month})")
        print(f"     Enrichissables ce mois  : {enrichissables} SIRETs")
        print(f"     SIRETs restants         : {n_no_tel - enrichissables}")
        print(f"     Mois nécessaires        : ~{mois} mois")
        print(f"  → RECOMMANDATION : enrichir les {enrichissables} lignes prioritaires")
        print(f"     (celles sans tél + les plus grandes villes en premier)")

    print()
    print(f"  Taux couverture Pappers estimé : ~20% (test empirique 10 SIRETs)")
    print(f"  → Téléphones attendus ce mois  : ~{tel_attendus} numéros supplémentaires")
    print()
    print("  ⛔ Enrichissement NON lancé — validation manuelle requise")
    print("=" * 65)
    print(f"\n  [OK] Fichier exporté : {OUTPUT_FILE}")
    print(f"       {total} lignes │ 📞 {n_tel} avec tél │ 🔲 {n_no_tel} sans tél\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  SCRAP MBT MULTI-VILLES — CAP Boulanger / Pâtissier")
    print(f"  LBA : ROME {', '.join(ROME_CODES)}")
    print(f"  LBB : NAF  {', '.join(NAF_CODES_LBB)}  (fallback ROME si non supporté)")
    print(f"  Villes : {len(CITIES)}  |  Max {MAX_PER_CITY} lignes/ville")
    print("=" * 65)

    # Obtenir le token LBB une seule fois (réutilisé pour toutes les villes)
    print("\n── Authentification LBB ─────────────────────────────────────")
    lbb_token = _get_lbb_token()

    # Dédup globale — SIRET → première ville qui le réclame
    global_sirets: set[str] = set()
    city_data: dict[str, list[dict]] = {}

    for city_name, cp, lat, lon, dist_km in CITIES:
        print(f"\n{'─'*65}")
        print(f"  📍 {city_name} ({cp}) — rayon {dist_km} km")
        print(f"{'─'*65}")

        # ── LBB ──
        lbb_rows: list[dict] = []
        if lbb_token:
            lbb_rows = fetch_lbb_city(lbb_token, city_name, lat, lon, dist_km)
            print(f"  [LBB] {len(lbb_rows)} entreprises brutes")
        else:
            print("  [LBB] Token indisponible — skippé")

        # ── LBA ──
        print(f"  [LBA] Requêtes ROME…")
        lba_rows = fetch_lba_city(city_name, lat, lon, dist_km)
        print(f"  [LBA] {len(lba_rows)} entreprises brutes")

        # ── Fusion locale ──
        by_siret = merge_local(lbb_rows, lba_rows)
        print(f"  Uniques locaux : {len(by_siret)}", end="")

        # ── Dédup globale ──
        # Retirer les SIRETs déjà assignés à une ville précédente
        available = {s: r for s, r in by_siret.items() if s not in global_sirets}
        print(f"  →  {len(available)} après dédup globale")

        # ── Tri : avec tél en premier, puis offre active, puis entreprise cible ──
        def _sort_key(r: dict) -> tuple:
            has_phone  = 0 if str(r.get("Téléphone", "")).strip() else 1
            type_order = {"Offre active": 0, "Entreprise cible": 1}
            return (has_phone, type_order.get(r.get("Type", ""), 2))

        sorted_rows = sorted(available.values(), key=_sort_key)

        # ── Cap 300 par ville ──
        selected = sorted_rows[:MAX_PER_CITY]

        # Enregistrer les SIRETs dans la dédup globale
        for row in selected:
            global_sirets.add(row["SIRET"])

        n_tel_city = sum(1 for r in selected if str(r.get("Téléphone", "")).strip())
        print(f"  Sélectionnés   : {len(selected)}  (📞 {n_tel_city} avec tél / "
              f"{len(selected)-n_tel_city} sans tél)")

        city_data[city_name] = selected

    print(f"\n{'─'*65}")
    print(f"  Total global : {sum(len(r) for r in city_data.values())} lignes")

    if not any(city_data.values()):
        print("[WARN] Aucune entreprise collectée.")
        sys.exit(0)

    # ── Export Excel ──
    export_excel(city_data, OUTPUT_FILE)

    # ── Bilan ──
    print_rapport(city_data)


if __name__ == "__main__":
    main()
