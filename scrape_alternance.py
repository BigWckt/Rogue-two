"""
Scraper d'entreprises en alternance/apprentissage
Sources : La Bonne Boite (LBB) + La Bonne Alternance (LBA) + Offres d'emploi FT (OFT)
Zone : 50km autour de Paris 1er (48.8603, 2.3477)
LBB  : ROME D1102/D1103 — entreprises à fort potentiel d'embauche (sans téléphone)
LBA  : ROME D1102/D1103 — offres alternance + recruteurs (avec téléphone)
OFT  : ROME D1102/D1103/D1104/D1101 — offres d'emploi FT avec contact.telephone

Notes LBB v2 (confirmé par documentation officielle francetravail.io) :
  - Scopes requis simultanément : search office api_labonneboitev2
  - Endpoint : /partenaire/labonneboite/v2/recherche
  - OAuth URL : authentification-partenaire.francetravail.io
"""

import re as _re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import sys

# ── Configuration ────────────────────────────────────────────────────────────

LATITUDE   = 48.8603
LONGITUDE  = 2.3477
DISTANCE_KM = 50
# Codes ROME étendus pour couvrir boulangerie/pâtisserie + secteurs connexes
# afin de collecter 150 entreprises ayant un numéro de téléphone visible
ROME_CODES = ["D1102", "D1103"]
# Codes ROME pour Offres d'emploi FT (plus large pour maximiser les téléphones)
ROME_CODES_OFT = ["D1102", "D1103", "D1104", "D1101"]
# Départements Île-de-France (50km autour de Paris)
DEPTS_IDF  = ["75", "77", "78", "91", "92", "93", "94", "95"]
# Codes NAF boulangerie-pâtisserie pour LBB (pas de filtre ROME disponible)
NAF_CODES  = ["10.71A", "10.71B", "10.71C", "10.71D"]
MAX_TOTAL  = 150
PHONE_ONLY = True   # n'exporter que les entreprises ayant un téléphone

# ── LBB (La Bonne Boite) ─────────────────────────────────────────────────────
# v2 — scopes et endpoint confirmés par support France Travail
LBB_ENDPOINT   = "https://api.francetravail.io/partenaire/labonneboite/v2/recherche"
# OAuth2 URL officielle (documentation francetravail.io)
LBB_OAUTH_URL         = "https://authentification-partenaire.francetravail.io/connexion/oauth2/access_token?realm=/partenaire"
# Fallback ProxyProConnect
LBB_OAUTH_URL_LEGACY  = "https://entreprise.francetravail.fr/connexion/oauth2/access_token?realm=/partenaire"
LBB_CLIENT_ID     = "PAR_claudecode_dbfd12ec4f6fe1174e46c36b762d98130ae05b4c33d069c1a5bebebe8573f33a"
LBB_CLIENT_SECRET = "a866591ed795a34b62c4d379e7f0cff3e393a59da7fbf3dfb04f101b90de2dd3"
# Scopes requis simultanément (confirmé support FT) — fallback avec application_id si nécessaire
LBB_SCOPE_CANDIDATES = [
    "search office api_labonneboitev2",
    f"search office api_labonneboitev2 application_{LBB_CLIENT_ID}",
]

# ── OFT (Offres d'emploi France Travail) ─────────────────────────────────────
OFT_ENDPOINT = "https://api.francetravail.io/partenaire/offresdemploi/v2/offres/search"
OFT_SCOPE    = "api_offresdemploiv2 o2dsoffre"
# Même client_id/secret que LBB (même application FT)
OFT_CLIENT_ID     = LBB_CLIENT_ID
OFT_CLIENT_SECRET = LBB_CLIENT_SECRET

# ── LBA (La Bonne Alternance) ─────────────────────────────────────────────────
# Note : /fr/explorer/recherche-offre est l'explorateur UI.
#        L'endpoint REST actif est /api/job/v1/search (même domaine).
LBA_ENDPOINT = "https://api.apprentissage.beta.gouv.fr/api/job/v1/search"
LBA_TOKEN = (
    "eyJhbGciOiJIUzI1NiJ9."
    "eyJfaWQiOiI2OGM4MzJkZDgxZGY5MmFiYTc2MDNhNzAiLCJhcGlfa2V5IjoieVFaYkpiZElCN1VydDNvb2"
    "w3aTRqN0lSRUhSQ25KSk5ya0djclpxZ1E0bz0iLCJvcmdhbmlzYXRpb24iOiJTa2lsbGFuZHlvdSIsImVt"
    "YWlsIjoiam9mZnJleS5sYWRtaXJhdWx0QHNraWxsYW5keW91LmNvbSIsImlzcyI6ImFwaSIsImlhdCI6MTc"
    "3MzMyMzg1NSwiZXhwIjoxNzg5NDg2Njk2fQ."
    "5P4D4gmCTezPIvycQEtKnPZBjPvcx8BHG1bF8r_Z4ts"
)

OUTPUT_FILE = "entreprises_alternance_paris.xlsx"

EXCEL_COLUMNS = [
    "Source",
    "Type",
    "SIRET",
    "Raison sociale",
    "Adresse",
    "Code postal",
    "Ville",
    "Téléphone",
    "Email",
    "Site web",
    "Code NAF",
    "Code ROME",
    "Distance Paris 1er (km)",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_safe(d, *keys, default=None):
    for k in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(k, default)
        if d is None:
            return default
    return d


def http_get(url, params=None, headers=None, timeout=20):
    """GET avec gestion des erreurs HTTP courantes."""
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        if resp.status_code == 429:
            print("  [WARN] 429 Too Many Requests — pause 10s puis réessai…")
            time.sleep(10)
            resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        if resp.status_code == 401:
            print(f"  [ERROR] 401 Unauthorized — vérifier les credentials. URL: {url}")
            return None
        if resp.status_code == 403:
            print(f"  [ERROR] 403 Forbidden (insufficient_scope) — URL: {url}")
            print("  [HINT]  Souscription LBB non active sur https://francetravail.io")
            return None
        resp.raise_for_status()
        # Rejeter les réponses HTML (page UI au lieu d'API REST)
        ct = resp.headers.get("content-type", "")
        if "html" in ct:
            print(f"  [ERROR] Réponse HTML reçue au lieu de JSON — URL incorrecte ? {url}")
            return None
        return resp.json()
    except requests.exceptions.Timeout:
        print(f"  [ERROR] Timeout après {timeout}s sur {url}")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"  [ERROR] Connexion impossible : {e}")
        return None
    except requests.exceptions.HTTPError as e:
        print(f"  [ERROR] HTTP {e.response.status_code} sur {url}")
        return None
    except ValueError:
        print(f"  [ERROR] Réponse non-JSON sur {url}")
        return None


# ── La Bonne Boite ────────────────────────────────────────────────────────────

_lbb_token_cache: dict = {}


def _get_lbb_token() -> str | None:
    """
    Obtient un Bearer token OAuth2 pour l'API LBB v2.
    Teste les scopes dans LBB_SCOPE_CANDIDATES et les deux URLs OAuth
    (ProxyProConnect en priorité, legacy en fallback) jusqu'au premier succès,
    puis valide que le token donne accès à l'API.
    """
    import time as _time
    now = _time.time()
    if _lbb_token_cache.get("token") and _lbb_token_cache.get("expires_at", 0) > now + 30:
        return _lbb_token_cache["token"]

    print("  [LBB] Recherche du token OAuth2 (v2)…")
    oauth_urls = [
        ("ProxyProConnect", LBB_OAUTH_URL),
        ("Legacy", LBB_OAUTH_URL_LEGACY),
    ]

    for scope in LBB_SCOPE_CANDIDATES:
        short_scope = scope[:60] + "…" if len(scope) > 60 else scope
        for url_label, oauth_url in oauth_urls:
            print(f"  [LBB] Test scope='{short_scope}' via {url_label}…")
            try:
                resp = requests.post(
                    oauth_url,
                    data={
                        "grant_type": "client_credentials",
                        "client_id": LBB_CLIENT_ID,
                        "client_secret": LBB_CLIENT_SECRET,
                        "scope": scope,
                    },
                    timeout=15,
                )
                if resp.status_code == 400:
                    err = resp.json().get("error_description", "?")
                    print(f"    → 400 Bad Request : {err}")
                    continue
                if resp.status_code == 401:
                    print("    → 401 Unauthorized (credentials invalides)")
                    # credentials invalides : inutile de tester d'autres URLs/scopes
                    return None
                resp.raise_for_status()
                token = resp.json().get("access_token")
                expires_in = int(resp.json().get("expires_in", 1800))
                print(f"    → Token obtenu (expire dans {expires_in}s) — test sur l'API…")

                # Valider que ce token donne vraiment accès à l'endpoint v2
                probe = requests.get(
                    LBB_ENDPOINT,
                    params={"latitude": LATITUDE, "longitude": LONGITUDE,
                            "distance": 1, "rome": "D1102", "page_size": 1},
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=10,
                )
                if probe.status_code == 200:
                    print(f"    → ✅ Token valide (scope='{short_scope}', url={url_label})")
                    _lbb_token_cache["token"] = token
                    _lbb_token_cache["expires_at"] = now + expires_in
                    return token
                else:
                    ct = probe.headers.get("content-type", "")
                    body = probe.text[:200] if "html" not in ct else "<HTML>"
                    print(f"    → API répond {probe.status_code} : {body}")
            except Exception as e:
                print(f"    → Erreur : {e}")
                continue

    print("  [LBB][ERROR] Aucune combinaison scope/URL n'a donné accès à l'API LBB v2.")
    print("  [HINT] Vérifier la souscription 'La Bonne Boite v2' sur https://francetravail.io")
    return None


def fetch_lbb() -> list[dict]:
    """
    Interroge l'API LBB v2 (/v2/recherche) par codes ROME.
    Paramètres confirmés : rome, latitude, longitude, distance, page, page_size.
    Note : l'API retourne items[] avec siret, company_name, naf, city, postcode,
           hiring_potential — mais PAS de téléphone (email = "yes"/"no" seulement).
    """
    token = _get_lbb_token()
    if not token:
        print("  [LBB] Token indisponible — skip")
        return []

    headers = {"Authorization": f"Bearer {token}"}
    results = []
    seen_sirets: set[str] = set()

    for rome_code in ROME_CODES:
        page = 1
        page_size = 100
        while True:
            params = {
                "latitude": LATITUDE,
                "longitude": LONGITUDE,
                "distance": DISTANCE_KM,
                "rome": rome_code,
                "page": page,
                "page_size": page_size,
            }
            print(f"\n[LBB] Requête ROME={rome_code} page={page} …")
            data = http_get(LBB_ENDPOINT, params=params, headers=headers)
            if data is None:
                break

            items = data.get("items") or []
            total_hits = data.get("hits", 0)
            print(f"  [LBB] {len(items)} item(s) (hits total: {total_hits})")

            for c in items:
                siret = str(c.get("siret") or "").strip()
                if not siret or siret in seen_sirets:
                    continue
                seen_sirets.add(siret)
                results.append({
                    "Source": "LBB",
                    "Type": "Entreprise cible",
                    "SIRET": siret,
                    "Raison sociale": c.get("company_name") or c.get("office_name") or "",
                    "Adresse": "",
                    "Code postal": str(c.get("postcode") or ""),
                    "Ville": c.get("city") or "",
                    "Téléphone": "",  # /v2/recherche ne fournit pas le téléphone
                    "Email": "",      # email="yes"/"no" — pas l'adresse réelle
                    "Site web": "",
                    "Code NAF": c.get("naf") or "",
                    "Code ROME": rome_code,
                    "Distance Paris 1er (km)": "",
                })

            if len(items) < page_size or len(results) >= MAX_TOTAL * 3:
                break
            page += 1

    print(f"  [LBB] {len(results)} entreprise(s) brute(s) au total")
    return results


# ── La Bonne Alternance ───────────────────────────────────────────────────────

def _parse_address(full_address: str) -> tuple[str, str, str]:
    """
    Parse '2 RUE CATULLE MENDES 75017 PARIS' → (rue, code_postal, ville).
    """
    if not full_address:
        return "", "", ""
    m = _re.search(r'(\d{5})\s+(.+)$', full_address.strip())
    if m:
        return (
            full_address[:m.start()].strip().rstrip(',').strip(),
            m.group(1),
            m.group(2).strip(),
        )
    return full_address, "", ""


def _extract_company_lba(item: dict, rome_code: str) -> dict | None:
    """Extrait les données entreprise d'un objet job ou recruiter LBA."""
    wp    = item.get("workplace") or {}
    apply = item.get("apply") or {}
    naf   = (wp.get("domain") or {}).get("naf") or {}

    siret = str(wp.get("siret") or "").strip()
    if not siret:
        return None

    rue, code_postal, ville = _parse_address(
        (wp.get("location") or {}).get("address") or ""
    )

    offer = item.get("offer") or {}
    rome  = (offer.get("rome_codes") or [rome_code])[0]

    return {
        "Source": "LBA",
        "Type": "Offre active",
        "SIRET": siret,
        "Raison sociale": wp.get("name") or wp.get("legal_name") or wp.get("brand") or "",
        "Adresse": rue,
        "Code postal": code_postal,
        "Ville": ville,
        "Téléphone": apply.get("phone") or "",
        "Email": "",
        "Site web": wp.get("website") or "",
        "Code NAF": naf.get("code") or "",
        "Code ROME": rome,
        "Distance Paris 1er (km)": "",
    }


def fetch_lba(rome_code: str) -> list[dict]:
    """
    Interroge l'API LBA pour un code ROME.
    - jobs[]       → Type "Offre active"      (offre postée, filtrée par ROME)
    - recruiters[] → Type "Entreprise cible"  (potentiel d'embauche, tous secteurs)
    Si un SIRET apparaît dans les deux, "Offre active" prend la priorité.
    Note : les recruiters ne sont pas filtrés par NAF/ROME côté API LBA —
           ils couvrent tous secteurs ayant recruté en alternance dans la zone.
    """
    params  = {"romes": rome_code, "latitude": LATITUDE,
               "longitude": LONGITUDE, "radius": DISTANCE_KM}
    headers = {"Authorization": f"Bearer {LBA_TOKEN}"}

    print(f"\n[LBA] Requête ROME={rome_code} …")
    data = http_get(LBA_ENDPOINT, params=params, headers=headers)
    if data is None:
        print(f"  [LBA] Aucune donnée reçue pour ROME={rome_code}")
        return []

    jobs       = data.get("jobs") or []
    recruiters = data.get("recruiters") or []
    print(f"  [LBA] {len(jobs)} offre(s) + {len(recruiters)} recruteur(s) bruts pour ROME={rome_code}")

    seen: dict[str, dict] = {}

    # D'abord les recruiters (priorité basse)
    for item in recruiters:
        company = _extract_company_lba(item, rome_code)
        if company is None:
            continue
        company["Type"] = "Entreprise cible"
        if company["SIRET"] not in seen:
            seen[company["SIRET"]] = company

    # Ensuite les jobs (écrasent les recruiters si même SIRET)
    for item in jobs:
        company = _extract_company_lba(item, rome_code)
        if company is None:
            continue
        company["Type"] = "Offre active"
        seen[company["SIRET"]] = company  # priorité absolue

    nb_offre  = sum(1 for c in seen.values() if c["Type"] == "Offre active")
    nb_cible  = sum(1 for c in seen.values() if c["Type"] == "Entreprise cible")
    print(f"  [LBA] {len(seen)} unique(s) — {nb_offre} offres actives, {nb_cible} entreprises cibles")
    return list(seen.values())


# ── Offres d'emploi France Travail (OFT) ─────────────────────────────────────

_oft_token_cache: dict = {}


def _get_oft_token() -> str | None:
    import time as _time
    now = _time.time()
    if _oft_token_cache.get("token") and _oft_token_cache.get("expires_at", 0) > now + 30:
        return _oft_token_cache["token"]
    resp = requests.post(
        LBB_OAUTH_URL,
        data={
            "grant_type": "client_credentials",
            "client_id": OFT_CLIENT_ID,
            "client_secret": OFT_CLIENT_SECRET,
            "scope": OFT_SCOPE,
        },
        timeout=15,
    )
    if resp.status_code != 200:
        print(f"  [OFT] Token KO {resp.status_code}: {resp.text[:100]}")
        return None
    token = resp.json().get("access_token")
    expires_in = int(resp.json().get("expires_in", 1800))
    _oft_token_cache["token"] = token
    _oft_token_cache["expires_at"] = now + expires_in
    return token


def fetch_oft() -> list[dict]:
    """
    Interroge l'API Offres d'emploi v2 (scope api_offresdemploiv2 o2dsoffre).
    Retourne les offres ayant contact.telephone renseigné.
    Pas de SIRET disponible — dédup par (nom_entreprise, code_postal).
    """
    token = _get_oft_token()
    if not token:
        print("  [OFT] Token indisponible — skip")
        return []

    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    results = []
    seen: set[str] = set()

    for dept in DEPTS_IDF:
        for rome in ROME_CODES_OFT:
            params = {"codeROME": rome, "departement": dept, "range": "0-149"}
            data = http_get(OFT_ENDPOINT, params=params, headers=headers)
            if data is None:
                continue
            offres = data.get("resultats") or []
            for o in offres:
                phone = (o.get("contact") or {}).get("telephone", "").strip()
                if not phone:
                    continue
                lieu   = o.get("lieuTravail") or {}
                entrep = o.get("entreprise") or {}
                nom    = entrep.get("nom") or ""
                cp     = lieu.get("codePostal") or ""
                key    = f"{nom.lower().strip()}|{cp}"
                if key in seen:
                    continue
                seen.add(key)
                results.append({
                    "Source": "OFT",
                    "Type": "Offre active",
                    "SIRET": "",
                    "Raison sociale": nom,
                    "Adresse": "",
                    "Code postal": cp,
                    "Ville": lieu.get("libelle", "").split(" - ")[-1].title(),
                    "Téléphone": phone,
                    "Email": (o.get("contact") or {}).get("courriel") or "",
                    "Site web": (o.get("contact") or {}).get("urlPostulation") or "",
                    "Code NAF": o.get("codeNAF") or "",
                    "Code ROME": o.get("romeCode") or rome,
                    "Distance Paris 1er (km)": "",
                })

    print(f"  [OFT] {len(results)} offre(s) avec téléphone")
    return results


# ── Déduplication et fusion des sources ──────────────────────────────────────

def merge_and_deduplicate(
    lbb_rows: list[dict], lba_rows: list[dict], oft_rows: list[dict]
) -> list[dict]:
    """
    Fusionne LBB, LBA et OFT.
    - Dédup LBB/LBA par SIRET.
    - OFT n'a pas de SIRET : dédup par (nom_entreprise, code_postal),
      puis enrichit un SIRET existant si le nom correspond.
    """
    def _dedup_by_siret(rows: list[dict]) -> dict[str, dict]:
        by_siret: dict[str, dict] = {}
        for row in rows:
            s = row["SIRET"]
            if not s:
                continue
            if s not in by_siret:
                by_siret[s] = row.copy()
            else:
                if row.get("Type") == "Offre active":
                    by_siret[s]["Type"] = "Offre active"
                for col in EXCEL_COLUMNS:
                    if not by_siret[s].get(col) and row.get(col):
                        by_siret[s][col] = row[col]
        return by_siret

    lbb_by_siret = _dedup_by_siret(lbb_rows)
    lba_by_siret = _dedup_by_siret(lba_rows)

    # Fusionner LBB + LBA par SIRET
    by_siret: dict[str, dict] = dict(lbb_by_siret)
    for s, row in lba_by_siret.items():
        if s not in by_siret:
            by_siret[s] = row.copy()
        else:
            by_siret[s]["Source"] = "LBB + LBA"
            by_siret[s]["Type"]   = "Offre active"
            for col in EXCEL_COLUMNS:
                if not by_siret[s].get(col) and row.get(col):
                    by_siret[s][col] = row[col]

    # Index nom→siret pour enrichissement OFT
    name_to_siret: dict[str, str] = {
        row["Raison sociale"].lower().strip(): s
        for s, row in by_siret.items()
        if row.get("Raison sociale")
    }

    oft_only: list[dict] = []
    for row in oft_rows:
        nom_key = row.get("Raison sociale", "").lower().strip()
        siret = name_to_siret.get(nom_key)
        if siret:
            # Enrichir l'entrée existante avec le téléphone OFT
            for col in EXCEL_COLUMNS:
                if not by_siret[siret].get(col) and row.get(col):
                    by_siret[siret][col] = row[col]
            if "OFT" not in by_siret[siret]["Source"]:
                by_siret[siret]["Source"] += " + OFT"
        else:
            oft_only.append(row.copy())

    return list(by_siret.values()) + oft_only


# ── Export Excel ──────────────────────────────────────────────────────────────

def export_excel(rows: list[dict], filepath: str) -> None:
    # Trier : "Offre active" en premier, puis "Entreprise cible"
    type_order = {"Offre active": 0, "Entreprise cible": 1}
    rows = sorted(rows, key=lambda r: type_order.get(r.get("Type", ""), 2))

    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    df = df.head(MAX_TOTAL)

    total        = len(df)
    nb_lbb       = (df["Source"] == "LBB").sum()
    nb_lba       = (df["Source"] == "LBA").sum()
    nb_oft       = (df["Source"] == "OFT").sum()
    nb_both      = df["Source"].str.contains(r"\+").sum()
    nb_offre     = (df["Type"] == "Offre active").sum()
    nb_cible     = (df["Type"] == "Entreprise cible").sum()
    nb_no_phone  = df["Téléphone"].apply(lambda x: not str(x).strip()).sum()

    print(f"\n── Synthèse ──────────────────────────────────────")
    print(f"  Total entreprises exportées : {total}")
    print(f"  LBB uniquement             : {nb_lbb}")
    print(f"  LBA uniquement             : {nb_lba}")
    print(f"  OFT uniquement             : {nb_oft}")
    print(f"  Multi-sources (LBB+LBA…)   : {nb_both}")
    print(f"  Offre active               : {nb_offre}")
    print(f"  Entreprise cible           : {nb_cible}")
    print(f"  Lignes sans téléphone      : {nb_no_phone}")
    print(f"──────────────────────────────────────────────────\n")

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Résultats_Scrap", index=False)

    wb = load_workbook(filepath)
    ws = wb["Résultats_Scrap"]

    header_fill  = PatternFill("solid", fgColor="0066FF")
    alt_fill     = PatternFill("solid", fgColor="E8F0FF")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    summary_fill = PatternFill("solid", fgColor="002080")
    header_font  = Font(bold=True, color="FFFFFF")
    summary_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = white_fill if row_idx % 2 == 0 else alt_fill
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")

    summary_row = ws.max_row + 2
    for i, (label, value) in enumerate([
        ("Total entreprises", total),
        ("dont LBB",          nb_lbb),
        ("dont LBA",          nb_lba),
        ("dont OFT",          nb_oft),
        ("dont multi-sources",nb_both),
        ("Offre active",      nb_offre),
        ("Entreprise cible",  nb_cible),
        ("Sans téléphone",    nb_no_phone),
    ]):
        for col_idx, val in enumerate([label, value], start=1):
            cell = ws.cell(row=summary_row + i, column=col_idx, value=val)
            cell.fill      = summary_fill
            cell.font      = summary_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"[OK] Fichier Excel exporté : {filepath}")


# ── Point d'entrée ────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  SCRAPER ALTERNANCE — Paris 1er, 50 km")
    print("  LBB v2 : ROME D1102/D1103  |  scope: search office api_labonneboitev2")
    print("  LBA    : ROME D1102/D1103")
    print("  OFT    : ROME D1102/D1103/D1104/D1101 — Offres d'emploi FT")
    print("  Mode   : 150 contacts avec téléphone uniquement")
    print("=" * 60)

    # ── LBB ──
    print("\n── La Bonne Boite (LBB) ──────────────────────────")
    all_lbb = fetch_lbb()

    # ── LBA ──
    print("\n── La Bonne Alternance (LBA) ─────────────────────")
    all_lba: list[dict] = []
    for rome in ROME_CODES:
        all_lba.extend(fetch_lba(rome))

    # ── OFT ──
    print("\n── Offres d'emploi France Travail (OFT) ──────────")
    all_oft = fetch_oft()

    print(f"\n[INFO] LBB total brut : {len(all_lbb)}")
    print(f"[INFO] LBA total brut : {len(all_lba)}")
    print(f"[INFO] OFT total brut : {len(all_oft)}")

    merged = merge_and_deduplicate(all_lbb, all_lba, all_oft)
    print(f"[INFO] Entreprises uniques après dédup : {len(merged)}")

    if not merged:
        print("[WARN] Aucune entreprise trouvée.")
        sys.exit(0)

    no_phone = sum(1 for r in merged if not str(r.get("Téléphone", "")).strip())
    print(f"[INFO] {no_phone}/{len(merged)} entreprise(s) sans téléphone")

    if PHONE_ONLY:
        merged = [r for r in merged if str(r.get("Téléphone", "")).strip()]
        print(f"[INFO] Filtrage PHONE_ONLY → {len(merged)} entreprise(s) avec téléphone")
        if len(merged) < MAX_TOTAL:
            print(f"[WARN] Seulement {len(merged)} entrées avec téléphone (objectif : {MAX_TOTAL})")

    export_excel(merged, OUTPUT_FILE)


if __name__ == "__main__":
    main()
