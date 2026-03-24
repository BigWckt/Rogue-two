"""
Scrap complet MBT (Métiers de Bouche) — LBA + LBB
===========================================================
LBA : ROME D1101, D1102, D1103, D1104, D1106, G1603
LBB : NAF  10.71A, 10.71B, 10.71C, 10.71D (+ fallback ROME boulangerie)
Zone: 50 km autour de Paris 1er (48.8603, 2.3477)

Règles :
  - Pas de limite de lignes — maximum disponible
  - Toutes les lignes conservées y compris sans téléphone
  - SIRET normalisé : str(int(float(siret))).zfill(14)
  - Déduplication par SIRET sur l'ensemble du fichier final

Étape 2 — rapport console :
  - Nombre total / avec téléphone / sans téléphone
  - Estimation coût Pappers (× 4 crédits) vs quota mensuel (2 000 crédits)
"""

import re as _re
import sys
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paramètres géographiques ──────────────────────────────────────────────────

LATITUDE    = 48.8603
LONGITUDE   = 2.3477
DISTANCE_KM = 50

# ── Codes secteur MBT ─────────────────────────────────────────────────────────

ROME_CODES_LBA = ["D1101", "D1102", "D1103", "D1104", "D1106", "G1603"]
NAF_CODES_LBB  = ["10.71A", "10.71B", "10.71C", "10.71D"]
# Fallback ROME si le filtre NAF n'est pas supporté par l'API LBB
ROME_CODES_LBB_FALLBACK = ["D1101", "D1102", "D1103", "D1104", "D1106", "G1603"]
DEPTS_IDF    = ["75", "77", "78", "91", "92", "93", "94", "95"]

# ── Authentification ──────────────────────────────────────────────────────────

LBB_ENDPOINT  = "https://api.francetravail.io/partenaire/labonneboite/v2/recherche"
LBB_OAUTH_URL = (
    "https://authentification-partenaire.francetravail.io"
    "/connexion/oauth2/access_token?realm=/partenaire"
)
LBB_OAUTH_URL_LEGACY = (
    "https://entreprise.francetravail.fr"
    "/connexion/oauth2/access_token?realm=/partenaire"
)
LBB_CLIENT_ID     = "PAR_claudecode_dbfd12ec4f6fe1174e46c36b762d98130ae05b4c33d069c1a5bebebe8573f33a"
LBB_CLIENT_SECRET = "a866591ed795a34b62c4d379e7f0cff3e393a59da7fbf3dfb04f101b90de2dd3"
LBB_SCOPE_CANDIDATES = [
    "search office api_labonneboitev2",
    f"search office api_labonneboitev2 application_{LBB_CLIENT_ID}",
]

LBA_ENDPOINT = "https://api.apprentissage.beta.gouv.fr/api/job/v1/search"
LBA_TOKEN = (
    "eyJhbGciOiJIUzI1NiJ9."
    "eyJfaWQiOiI2OGM4MzJkZDgxZGY5MmFiYTc2MDNhNzAiLCJhcGlfa2V5IjoieVFaYkpiZElCN1VydDNvb2"
    "w3aTRqN0lSRUhSQ25KSk5ya0djclpxZ1E0bz0iLCJvcmdhbmlzYXRpb24iOiJTa2lsbGFuZHlvdSIsImVt"
    "YWlsIjoiam9mZnJleS5sYWRtaXJhdWx0QHNraWxsYW5keW91LmNvbSIsImlzcyI6ImFwaSIsImlhdCI6MTc"
    "3MzMyMzg1NSwiZXhwIjoxNzg5NDg2Njk2fQ."
    "5P4D4gmCTezPIvycQEtKnPZBjPvcx8BHG1bF8r_Z4ts"
)

# ── Pappers (estimation uniquement — pas d'appels dans ce script) ─────────────

PAPPERS_CREDITS_PER_SIRET = 4
PAPPERS_MONTHLY_QUOTA     = 2_000

# ── Sortie ────────────────────────────────────────────────────────────────────

OUTPUT_FILE  = "mbt_entreprises.xlsx"
EXCEL_COLUMNS = [
    "Source", "Type", "SIRET", "Raison sociale",
    "Adresse", "Code postal", "Ville",
    "Téléphone", "Email",
    "Code NAF", "Code ROME", "Distance (km)",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_siret(raw) -> str:
    """Convertit n'importe quel format de SIRET (float, int, str) en str 14 chiffres."""
    if raw is None or str(raw).strip() in ("", "nan", "None"):
        return ""
    try:
        return str(int(float(str(raw).strip()))).zfill(14)
    except (ValueError, OverflowError):
        return str(raw).strip()


def http_get(url, params=None, headers=None, timeout=20):
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        if resp.status_code == 429:
            print("  [WARN] 429 Too Many Requests — pause 10s…")
            time.sleep(10)
            resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        if resp.status_code == 401:
            print(f"  [ERROR] 401 Unauthorized — {url}")
            return None
        if resp.status_code == 403:
            print(f"  [ERROR] 403 Forbidden — {url}")
            return None
        resp.raise_for_status()
        if "html" in resp.headers.get("content-type", ""):
            print(f"  [ERROR] Réponse HTML reçue (URL incorrecte ?) — {url}")
            return None
        return resp.json()
    except requests.exceptions.Timeout:
        print(f"  [ERROR] Timeout {timeout}s — {url}")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"  [ERROR] Connexion : {e}")
        return None
    except requests.exceptions.HTTPError as e:
        print(f"  [ERROR] HTTP {e.response.status_code} — {url}")
        return None
    except ValueError:
        print(f"  [ERROR] Réponse non-JSON — {url}")
        return None


def _parse_address(full_address: str) -> tuple[str, str, str]:
    if not full_address:
        return "", "", ""
    m = _re.search(r'(\d{5})\s+(.+)$', full_address.strip())
    if m:
        return (
            full_address[:m.start()].strip().rstrip(',').strip(),
            m.group(1),
            m.group(2).strip(),
        )
    return full_address, "", ""


# ── OAuth LBB ─────────────────────────────────────────────────────────────────

_lbb_token_cache: dict = {}


def _get_lbb_token() -> str | None:
    now = time.time()
    if _lbb_token_cache.get("token") and _lbb_token_cache.get("expires_at", 0) > now + 30:
        return _lbb_token_cache["token"]

    print("  [LBB] Obtention token OAuth2…")
    oauth_urls = [("ProxyProConnect", LBB_OAUTH_URL), ("Legacy", LBB_OAUTH_URL_LEGACY)]

    for scope in LBB_SCOPE_CANDIDATES:
        for label, oauth_url in oauth_urls:
            try:
                resp = requests.post(oauth_url, data={
                    "grant_type": "client_credentials",
                    "client_id": LBB_CLIENT_ID,
                    "client_secret": LBB_CLIENT_SECRET,
                    "scope": scope,
                }, timeout=15)
                if resp.status_code in (400, 401):
                    continue
                resp.raise_for_status()
                token = resp.json().get("access_token")
                expires_in = int(resp.json().get("expires_in", 1800))

                # Valider que le token donne accès à l'API
                probe = requests.get(LBB_ENDPOINT, params={
                    "latitude": LATITUDE, "longitude": LONGITUDE,
                    "distance": 1, "rome": "D1102", "page_size": 1,
                }, headers={"Authorization": f"Bearer {token}"}, timeout=10)

                if probe.status_code == 200:
                    print(f"  [LBB] ✅ Token valide (scope='{scope[:50]}', {label})")
                    _lbb_token_cache["token"] = token
                    _lbb_token_cache["expires_at"] = now + expires_in
                    return token
                else:
                    print(f"  [LBB] Token KO sur API : HTTP {probe.status_code}")
            except Exception as e:
                print(f"  [LBB] Erreur : {e}")
                continue

    print("  [LBB][ERROR] Aucune combinaison scope/URL valide — LBB skippé")
    return None


# ── LBB fetch ─────────────────────────────────────────────────────────────────

def _fetch_lbb_by_param(token: str, param_name: str, param_values: list[str]) -> list[dict]:
    """
    Collecte toutes les pages pour chaque valeur dans param_values.
    param_name : 'naf' ou 'rome'
    """
    headers  = {"Authorization": f"Bearer {token}"}
    results  = []
    seen_sirets: set[str] = set()

    for val in param_values:
        page     = 1
        page_size = 100
        total_hits_reported = None

        while True:
            params = {
                "latitude": LATITUDE,
                "longitude": LONGITUDE,
                "distance": DISTANCE_KM,
                param_name: val,
                "page": page,
                "page_size": page_size,
            }
            print(f"  [LBB] {param_name.upper()}={val} page={page}…", end=" ", flush=True)
            data = http_get(LBB_ENDPOINT, params=params, headers=headers)
            if data is None:
                print()
                break

            # Détecter si le param NAF est inconnu (API renvoi items vides + hits=0)
            items = data.get("items") or []
            hits  = data.get("hits", 0) or 0
            if total_hits_reported is None:
                total_hits_reported = hits
            print(f"{len(items)} items (total hits: {hits})")

            for c in items:
                siret = normalize_siret(c.get("siret") or "")
                if not siret or siret in seen_sirets:
                    continue
                seen_sirets.add(siret)
                results.append({
                    "Source": "LBB",
                    "Type": "Entreprise cible",
                    "SIRET": siret,
                    "Raison sociale": c.get("company_name") or c.get("office_name") or "",
                    "Adresse": "",
                    "Code postal": str(c.get("postcode") or ""),
                    "Ville": c.get("city") or "",
                    "Téléphone": "",
                    "Email": "",
                    "Code NAF": c.get("naf") or val if param_name == "naf" else c.get("naf") or "",
                    "Code ROME": val if param_name == "rome" else "",
                    "Distance (km)": "",
                })

            if len(items) < page_size:
                break
            page += 1
            time.sleep(0.3)

    return results


def fetch_lbb() -> list[dict]:
    """
    Tente d'abord un filtre NAF (10.71x).
    Si l'API renvoie 0 résultats sur tous les codes NAF, bascule sur ROME.
    """
    token = _get_lbb_token()
    if not token:
        return []

    print(f"\n  [LBB] Tentative filtre NAF : {NAF_CODES_LBB}")
    results_naf = _fetch_lbb_by_param(token, "naf", NAF_CODES_LBB)

    if results_naf:
        print(f"  [LBB] Filtre NAF → {len(results_naf)} entreprises")
        return results_naf

    print(f"  [LBB] Filtre NAF sans résultat — fallback ROME : {ROME_CODES_LBB_FALLBACK}")
    results_rome = _fetch_lbb_by_param(token, "rome", ROME_CODES_LBB_FALLBACK)
    print(f"  [LBB] Filtre ROME → {len(results_rome)} entreprises")
    return results_rome


# ── LBA fetch ─────────────────────────────────────────────────────────────────

def _extract_company_lba(item: dict, rome_code: str) -> dict | None:
    wp    = item.get("workplace") or {}
    apply = item.get("apply") or {}
    naf   = (wp.get("domain") or {}).get("naf") or {}

    raw_siret = wp.get("siret") or ""
    siret = normalize_siret(raw_siret)
    if not siret:
        return None

    rue, code_postal, ville = _parse_address(
        (wp.get("location") or {}).get("address") or ""
    )
    offer = item.get("offer") or {}
    rome  = (offer.get("rome_codes") or [rome_code])[0]

    return {
        "Source": "LBA",
        "Type": "Offre active",
        "SIRET": siret,
        "Raison sociale": wp.get("name") or wp.get("legal_name") or wp.get("brand") or "",
        "Adresse": rue,
        "Code postal": code_postal,
        "Ville": ville,
        "Téléphone": apply.get("phone") or "",
        "Email": "",
        "Code NAF": naf.get("code") or "",
        "Code ROME": rome,
        "Distance (km)": "",
    }


def fetch_lba(rome_code: str) -> list[dict]:
    params  = {
        "romes": rome_code,
        "latitude": LATITUDE,
        "longitude": LONGITUDE,
        "radius": DISTANCE_KM,
    }
    headers = {"Authorization": f"Bearer {LBA_TOKEN}"}

    print(f"  [LBA] ROME={rome_code}…", end=" ", flush=True)
    data = http_get(LBA_ENDPOINT, params=params, headers=headers)
    if data is None:
        print("aucune donnée")
        return []

    jobs       = data.get("jobs") or []
    recruiters = data.get("recruiters") or []
    print(f"{len(jobs)} offres + {len(recruiters)} recruteurs")

    seen: dict[str, dict] = {}

    for item in recruiters:
        company = _extract_company_lba(item, rome_code)
        if company is None:
            continue
        company["Type"] = "Entreprise cible"
        if company["SIRET"] not in seen:
            seen[company["SIRET"]] = company

    for item in jobs:
        company = _extract_company_lba(item, rome_code)
        if company is None:
            continue
        company["Type"] = "Offre active"
        seen[company["SIRET"]] = company  # priorité

    return list(seen.values())


# ── Fusion + déduplication ────────────────────────────────────────────────────

def merge_and_deduplicate(lbb_rows: list[dict], lba_rows: list[dict]) -> list[dict]:
    """
    Fusionne LBB + LBA par SIRET.
    Règles de priorité :
      - "Offre active" > "Entreprise cible"
      - LBA > LBB pour le téléphone (LBB n'en a jamais)
      - Champs manquants comblés depuis la seconde source
    """
    by_siret: dict[str, dict] = {}

    def _upsert(row: dict, priority_source: str) -> None:
        siret = row.get("SIRET", "")
        if not siret:
            return
        if siret not in by_siret:
            by_siret[siret] = row.copy()
            return
        existing = by_siret[siret]
        # Mettre à jour le libellé source
        sources = set(existing["Source"].split(" + "))
        sources.add(priority_source)
        existing["Source"] = " + ".join(sorted(sources))
        # Priorité au type "Offre active"
        if row.get("Type") == "Offre active":
            existing["Type"] = "Offre active"
        # Combler les champs vides
        for col in EXCEL_COLUMNS:
            if not str(existing.get(col, "")).strip() and str(row.get(col, "")).strip():
                existing[col] = row[col]

    # LBB en premier (priorité basse)
    for row in lbb_rows:
        _upsert(row, "LBB")

    # LBA ensuite (priorité haute — téléphone, offre active)
    for row in lba_rows:
        _upsert(row, "LBA")

    return list(by_siret.values())


# ── Export Excel ──────────────────────────────────────────────────────────────

def export_excel(rows: list[dict], filepath: str) -> None:
    # Tri : Offre active en premier, puis Entreprise cible ; dans chaque groupe, avec tél en premier
    def sort_key(r):
        type_order = {"Offre active": 0, "Entreprise cible": 1}
        has_phone  = 0 if str(r.get("Téléphone", "")).strip() else 1
        return (type_order.get(r.get("Type", ""), 2), has_phone)

    rows = sorted(rows, key=sort_key)
    df   = pd.DataFrame(rows, columns=EXCEL_COLUMNS)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="MBT_Entreprises", index=False)

    wb = load_workbook(filepath)
    ws = wb["MBT_Entreprises"]

    hdr_fill  = PatternFill("solid", fgColor="1A3A6B")
    ok_fill   = PatternFill("solid", fgColor="D4EDDA")
    ko_fill   = PatternFill("solid", fgColor="FFF3CD")
    alt_fill  = PatternFill("solid", fgColor="EBF3FF")
    white_fill= PatternFill("solid", fgColor="FFFFFF")

    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    tel_col = EXCEL_COLUMNS.index("Téléphone") + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        base = white_fill if row_idx % 2 == 0 else alt_fill
        for cell in row:
            cell.fill      = base
            cell.alignment = Alignment(vertical="center")
        tel_cell = ws.cell(row_idx, tel_col)
        if str(tel_cell.value or "").strip():
            tel_cell.fill = ok_fill
            tel_cell.font = Font(bold=True, color="1A7A1A")
        else:
            tel_cell.fill = ko_fill

    col_widths = [14, 16, 16, 38, 35, 12, 22, 16, 28, 12, 12, 14]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    wb.save(filepath)


# ── Rapport bilan + estimation Pappers ───────────────────────────────────────

def print_rapport(rows: list[dict]) -> None:
    total    = len(rows)
    n_tel    = sum(1 for r in rows if str(r.get("Téléphone", "")).strip())
    n_no_tel = total - n_tel

    n_lbb = sum(1 for r in rows if r.get("Source") == "LBB")
    n_lba = sum(1 for r in rows if r.get("Source") == "LBA")
    n_both= sum(1 for r in rows if "+" in str(r.get("Source", "")))
    n_offre = sum(1 for r in rows if r.get("Type") == "Offre active")
    n_cible = sum(1 for r in rows if r.get("Type") == "Entreprise cible")

    credits_needed  = n_no_tel * PAPPERS_CREDITS_PER_SIRET
    max_enrichissements = PAPPERS_MONTHLY_QUOTA // PAPPERS_CREDITS_PER_SIRET
    enrichissables  = min(n_no_tel, max_enrichissements)
    credits_used    = enrichissables * PAPPERS_CREDITS_PER_SIRET
    credits_remaining = PAPPERS_MONTHLY_QUOTA - credits_used

    # Estimation taux de couverture Pappers (basée sur le test empirique de 10 SIRETs : 2/10 = 20%)
    coverage_rate   = 0.20
    tel_attendus    = int(enrichissables * coverage_rate)

    print("\n" + "=" * 65)
    print("  BILAN DU SCRAP MBT")
    print("=" * 65)
    print(f"  Total lignes collectées        : {total}")
    print(f"  ├─ LBB uniquement              : {n_lbb}")
    print(f"  ├─ LBA uniquement              : {n_lba}")
    print(f"  └─ Multi-sources (LBB+LBA)     : {n_both}")
    print()
    print(f"  Offre active                   : {n_offre}")
    print(f"  Entreprise cible               : {n_cible}")
    print()
    print(f"  AVEC téléphone                 : {n_tel}  ({n_tel/total*100:.1f}%)")
    print(f"  SANS téléphone                 : {n_no_tel}  ({n_no_tel/total*100:.1f}%)")
    print()
    print("─" * 65)
    print("  ESTIMATION ENRICHISSEMENT PAPPERS")
    print("─" * 65)
    print(f"  Lignes sans téléphone          : {n_no_tel}")
    print(f"  Coût par SIRET                 : {PAPPERS_CREDITS_PER_SIRET} crédits")
    print(f"  Coût total si 100% enrichi     : {credits_needed:,} crédits")
    print()
    print(f"  Quota mensuel disponible       : {PAPPERS_MONTHLY_QUOTA:,} crédits")
    print(f"  → Max enrichissements/mois     : {max_enrichissements} SIRETs")
    print()
    if n_no_tel <= max_enrichissements:
        print(f"  ✅ Budget suffisant pour tout enrichir ({n_no_tel} ≤ {max_enrichissements})")
        print(f"     Crédits consommés   : {credits_needed:,} / {PAPPERS_MONTHLY_QUOTA:,}")
        print(f"     Crédits restants    : {PAPPERS_MONTHLY_QUOTA - credits_needed:,}")
    else:
        print(f"  ⚠️  Budget insuffisant pour tout enrichir")
        print(f"     Enrichissables ce mois   : {max_enrichissements} / {n_no_tel} SIRETs")
        print(f"     Crédits consommés        : {credits_used:,} / {PAPPERS_MONTHLY_QUOTA:,}")
        print(f"     SIRETs non traités       : {n_no_tel - max_enrichissements}")

    print()
    print(f"  Taux couverture Pappers estimé : ~{coverage_rate*100:.0f}% (basé sur test 10 SIRETs)")
    print(f"  → Téléphones attendus          : ~{tel_attendus} numéros supplémentaires")
    print()
    print(f"  ⛔ Enrichissement NON lancé — validation manuelle requise")
    print("=" * 65)
    print(f"\n  [OK] Fichier exporté : {OUTPUT_FILE}")
    print(f"       {total} lignes │ 📞 {n_tel} avec tél │ 🔲 {n_no_tel} sans tél\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  SCRAP COMPLET MBT — Métiers de Bouche")
    print(f"  Zone    : {DISTANCE_KM} km autour de Paris 1er ({LATITUDE}, {LONGITUDE})")
    print(f"  LBA     : ROME {', '.join(ROME_CODES_LBA)}")
    print(f"  LBB     : NAF  {', '.join(NAF_CODES_LBB)}")
    print("  Règles  : pas de limite, toutes lignes, dédup SIRET")
    print("=" * 65)

    # ── LBB ──────────────────────────────────────────────────────────────────
    print("\n── La Bonne Boite (LBB) ─────────────────────────────────────")
    all_lbb = fetch_lbb()
    print(f"  → {len(all_lbb)} entreprises LBB brutes")

    # ── LBA ──────────────────────────────────────────────────────────────────
    print("\n── La Bonne Alternance (LBA) ─────────────────────────────────")
    all_lba: list[dict] = []
    for rome in ROME_CODES_LBA:
        rows = fetch_lba(rome)
        all_lba.extend(rows)
    print(f"  → {len(all_lba)} entrées LBA brutes (avant dédup)")

    print(f"\n[INFO] LBB brut : {len(all_lbb)} | LBA brut : {len(all_lba)}")

    # ── Fusion ───────────────────────────────────────────────────────────────
    merged = merge_and_deduplicate(all_lbb, all_lba)
    print(f"[INFO] Entreprises uniques après fusion/dédup : {len(merged)}")

    if not merged:
        print("[WARN] Aucune entreprise collectée — vérifier les APIs.")
        sys.exit(0)

    # ── Export ───────────────────────────────────────────────────────────────
    export_excel(merged, OUTPUT_FILE)

    # ── Bilan ────────────────────────────────────────────────────────────────
    print_rapport(merged)


if __name__ == "__main__":
    main()
